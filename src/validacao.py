import logging
from pathlib import Path
import unicodedata

from openpyxl import load_workbook
import pandas as pd


COLUNAS_ESTRUTURA = [
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
]

COLUNAS_OBRIGATORIAS = COLUNAS_ESTRUTURA[:-1]

ABA_FORMULARIO_ANALISE = "Formulario_Analise"
TIPO_ERRO_RN02 = "Dados Obrigatorios Ausentes"
ACAO_RECOMENDADA_RN02 = "Preencher campos obrigatorios"


def carregar_planilha(caminho_arquivo):
    caminho = Path(caminho_arquivo)

    if caminho.suffix.lower() != ".xlsx":
        raise ValueError("O arquivo de entrada deve estar no formato .xlsx")

    df_bruto = pd.read_excel(caminho, header=None)
    linha_cabecalho = encontrar_linha_cabecalho(df_bruto)

    if linha_cabecalho is None:
        return pd.DataFrame()

    colunas = df_bruto.iloc[linha_cabecalho].astype(str).str.strip()
    df = df_bruto.iloc[linha_cabecalho + 1 :].copy()
    df.columns = colunas
    df = df.loc[:, ~df.columns.isin(["", "nan", "NaN"])]
    df = df.dropna(how="all")
    return filtrar_linhas_de_registro(df)


def encontrar_linha_cabecalho(df):
    colunas_estrutura = set(COLUNAS_ESTRUTURA)

    for indice, linha in df.iterrows():
        valores = {str(valor).strip() for valor in linha.dropna()}

        if colunas_estrutura.issubset(valores):
            return indice

    return None


def filtrar_linhas_de_registro(df):
    colunas_existentes = [coluna for coluna in COLUNAS_ESTRUTURA if coluna in df.columns]

    if not colunas_existentes:
        return df

    preenchidos = df[colunas_existentes].notna() & ~df[colunas_existentes].astype(str).apply(
        lambda coluna: coluna.str.strip().eq("")
    )
    minimo_campos = min(4, len(colunas_existentes))
    return df.loc[preenchidos.sum(axis=1) >= minimo_campos]


def valida_estrutura(caminho_arquivo="dados.xlsx"):
    """RN01: valida se o arquivo .xlsx possui exatamente as 8 colunas do PDD."""
    df = carregar_planilha(caminho_arquivo)
    faltantes = obter_colunas_faltantes(df)
    extras = obter_colunas_extras(df)

    if faltantes:
        logging.error("Colunas obrigatorias ausentes: %s", sorted(faltantes))
        return False

    if extras:
        logging.error("Colunas nao previstas no layout RN01: %s", sorted(extras))
        return False

    if list(df.columns) != COLUNAS_ESTRUTURA:
        logging.error("Ordem das colunas invalida para o layout RN01.")
        return False

    logging.info("As 8 colunas obrigatorias do layout RN01 estao presentes.")
    return True


def valida_campos_obrigatorios(caminho_arquivo="dados.xlsx"):
    """RN02: valida campos obrigatorios e aponta registros com dados ausentes."""
    df = carregar_planilha(caminho_arquivo)
    faltantes = obter_colunas_faltantes(df)

    if faltantes:
        logging.error("Colunas obrigatorias ausentes: %s", sorted(faltantes))
        return False

    erros = encontrar_erros_rn02(df)
    valido = not erros

    for erro in erros:
        logging.error(
            "Campos obrigatorios vazios na linha %s: %s",
            erro["linha"],
            ", ".join(erro["campos"]),
        )

    if valido:
        logging.info("Todos os campos obrigatorios estao preenchidos.")

    return valido


def registrar_erros_rn02(caminho_arquivo="dados.xlsx"):
    df = carregar_planilha(caminho_arquivo)
    erros = encontrar_erros_rn02(df)

    if not erros:
        return []

    workbook = load_workbook(caminho_arquivo)

    if ABA_FORMULARIO_ANALISE not in workbook.sheetnames:
        worksheet = workbook.create_sheet(ABA_FORMULARIO_ANALISE)
        criar_cabecalho_formulario(worksheet)
    else:
        worksheet = workbook[ABA_FORMULARIO_ANALISE]

    linha_cabecalho = encontrar_linha_cabecalho_formulario(worksheet)

    if linha_cabecalho is None:
        linha_cabecalho = worksheet.max_row + 1
        criar_cabecalho_formulario(worksheet, linha_cabecalho)

    colunas = mapear_colunas_formulario(worksheet, linha_cabecalho)
    linhas_existentes = mapear_linhas_rn02_existentes(worksheet, linha_cabecalho, colunas)
    proxima_linha = encontrar_proxima_linha_vazia(worksheet, linha_cabecalho)

    for erro in erros:
        linha_registro = linhas_existentes.get(erro["linha"], proxima_linha)
        worksheet.cell(linha_registro, colunas["linha"]).value = erro["linha"]
        worksheet.cell(linha_registro, colunas["lote_id"]).value = erro["lote_id"]
        worksheet.cell(linha_registro, colunas["tipo"]).value = TIPO_ERRO_RN02
        worksheet.cell(linha_registro, colunas["regra"]).value = "RN02"
        worksheet.cell(linha_registro, colunas["acao"]).value = (
            f"{ACAO_RECOMENDADA_RN02}: {', '.join(erro['campos'])}"
        )

        if linha_registro == proxima_linha:
            proxima_linha += 1

    workbook.save(caminho_arquivo)
    return erros


def encontrar_erros_rn02(df):
    erros = []

    for indice, registro in df.iterrows():
        campos_vazios = []

        for coluna in COLUNAS_OBRIGATORIAS:
            valor = registro[coluna]

            if pd.isna(valor) or str(valor).strip() == "":
                campos_vazios.append(coluna)

        if campos_vazios:
            erros.append(
                {
                    "linha": int(indice) + 1,
                    "lote_id": "" if pd.isna(registro["lote_id"]) else registro["lote_id"],
                    "campos": campos_vazios,
                }
            )

    return erros


def criar_cabecalho_formulario(worksheet, linha=1):
    cabecalhos = [
        "Linha",
        "lote_id",
        "Tipo de divergencia encontrada",
        "Regra(s) violada(s)",
        "Acao recomendada",
        "Confirmado no gabarito?",
    ]

    for coluna, cabecalho in enumerate(cabecalhos, start=1):
        worksheet.cell(linha, coluna).value = cabecalho


def encontrar_linha_cabecalho_formulario(worksheet):
    campos_necessarios = {"linha", "lote_id", "tipo", "regra", "acao"}

    for linha in range(1, worksheet.max_row + 1):
        valores = {
            classificar_coluna_formulario(worksheet.cell(linha, coluna).value)
            for coluna in range(1, worksheet.max_column + 1)
        }

        if campos_necessarios.issubset(valores):
            return linha

    return None


def mapear_colunas_formulario(worksheet, linha_cabecalho):
    colunas = {}

    for coluna in range(1, worksheet.max_column + 1):
        tipo = classificar_coluna_formulario(worksheet.cell(linha_cabecalho, coluna).value)

        if tipo:
            colunas[tipo] = coluna

    return colunas


def classificar_coluna_formulario(valor):
    texto = normalizar_texto(valor)

    if texto == "linha":
        return "linha"

    if texto == "lote_id":
        return "lote_id"

    if "tipo de divergencia" in texto:
        return "tipo"

    if "regra" in texto and "violada" in texto:
        return "regra"

    if "acao recomendada" in texto:
        return "acao"

    return None


def encontrar_proxima_linha_vazia(worksheet, linha_cabecalho):
    return worksheet.max_row + 1


def mapear_linhas_rn02_existentes(worksheet, linha_cabecalho, colunas):
    linhas = {}

    for linha in range(linha_cabecalho + 1, worksheet.max_row + 1):
        valor_linha = worksheet.cell(linha, colunas["linha"]).value
        valor_regra = worksheet.cell(linha, colunas["regra"]).value

        if "rn02" not in normalizar_texto(valor_regra):
            continue

        try:
            linhas[int(valor_linha)] = linha
        except (TypeError, ValueError):
            continue

    return linhas


def normalizar_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(caractere for caractere in texto if not unicodedata.combining(caractere))


def obter_colunas_faltantes(df):
    return set(COLUNAS_ESTRUTURA) - set(df.columns)


def obter_colunas_extras(df):
    return set(df.columns) - set(COLUNAS_ESTRUTURA)
