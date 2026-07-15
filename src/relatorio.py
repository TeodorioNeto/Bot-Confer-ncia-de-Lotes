"""
relatorio.py - Geração do relatório de divergências em .xlsx.

Recebe a lista de divergências encontradas pelo bot (RN01-RN07) e
gera uma planilha .xlsx formatada, pronta para envio ao time de
qualidade.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

CABECALHO = ["Linha", "Lote ID", "Regra Violada", "Problema Encontrado"]


def gerar_relatorio_divergencias(divergencias, caminho_saida=None):
    """
    Gera um arquivo .xlsx com as divergências encontradas na inspeção.

    Args:
        divergencias: lista de dicts, cada um com as chaves
                       'linha', 'lote_id', 'regra', 'problema'
                       (formato retornado por bot.processar_inspecao()).
        caminho_saida: caminho do arquivo .xlsx a ser gerado. Se None,
                        usa 'data/output/relatorio_divergencias_<data>.xlsx'.

    Returns:
        O caminho (Path) do arquivo gerado.
    """
    if caminho_saida is None:
        pasta_saida = Path("data/output")
        pasta_saida.mkdir(parents=True, exist_ok=True)
        data_hoje = datetime.now().strftime("%Y-%m-%d")
        caminho_saida = pasta_saida / f"relatorio_divergencias_{data_hoje}.xlsx"
    else:
        caminho_saida = Path(caminho_saida)
        caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Divergencias"

    ws.append(CABECALHO)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

    for divergencia in divergencias:
        ws.append(
            [
                divergencia.get("linha"),
                divergencia.get("lote_id"),
                divergencia.get("regra"),
                divergencia.get("problema"),
            ]
        )

    for coluna in ws.columns:
        maior_valor = max((len(str(celula.value)) for celula in coluna if celula.value), default=10)
        letra_coluna = coluna[0].column_letter
        ws.column_dimensions[letra_coluna].width = min(maior_valor + 2, 60)

    wb.save(caminho_saida)
    return caminho_saida