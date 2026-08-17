"""
src/web_automation_playwright.py
Orquestrador Playwright para a tela web simulada de inspecao de lotes.
"""

import os
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright

from config import ARQUIVO_INSPECAO
from src.base_referencia import carregar_base_referencia
from src.logger import setup_logger
from src.pages.form_page import FormPagePlaywright
from src.pages.login_page import LoginPagePlaywright
from src.validacao import COLUNAS_ESTRUTURA
from src.web_automation import analisar_dados_lote, iniciar_browser
from src.web_evidencias import montar_caminho_screenshot, obter_url_automacao

load_dotenv()

logger = setup_logger("PlaywrightEngine")


def emitir_log(mensagem, tipo="info", callback_log=None):
    if tipo == "warn":
        logger.warning(mensagem)
    elif tipo == "error":
        logger.error(mensagem)
    else:
        logger.info(mensagem)

    if callback_log:
        try:
            callback_log(mensagem, tipo)
        except Exception:
            pass


def carregar_datapool_tratado():
    caminho_planilha = Path(ARQUIVO_INSPECAO)
    lotes_tratados = []

    if caminho_planilha.exists():
        logger.info("Planilha carregada com sucesso: %s", caminho_planilha)
        workbook = openpyxl.load_workbook(caminho_planilha, read_only=True, data_only=True)
        try:
            for ws in workbook.worksheets:
                linha_cabecalho = _encontrar_linha_cabecalho(ws)
                if linha_cabecalho is None:
                    continue

                for numero_linha, linha in enumerate(
                    ws.iter_rows(
                        min_row=linha_cabecalho + 1,
                        max_col=len(COLUNAS_ESTRUTURA),
                        values_only=True,
                    ),
                    start=linha_cabecalho + 1,
                ):
                    if linha is None or all(valor is None for valor in linha):
                        break

                    preenchidos = sum(
                        valor is not None and bool(str(valor).strip())
                        for valor in linha
                    )
                    if preenchidos < 4:
                        continue

                    item = dict(zip(COLUNAS_ESTRUTURA, linha))
                    item["lote"] = item.get("lote_id")
                    item["linha_planilha"] = numero_linha
                    lotes_tratados.append(item)
        finally:
            workbook.close()
    else:
        logger.warning(
            "Planilha de entrada nao encontrada em '%s'. Nenhum lote sera processado.",
            caminho_planilha,
        )
        return []

    return lotes_tratados


def _encontrar_linha_cabecalho(ws):
    for numero_linha in range(1, ws.max_row + 1):
        valores = [
            ws.cell(numero_linha, coluna).value
            for coluna in range(1, len(COLUNAS_ESTRUTURA) + 1)
        ]
        if valores == COLUNAS_ESTRUTURA:
            return numero_linha
    return None


def processar_item_playwright(dados_lote, delay_passo=0, callback_log=None, theme="dark"):
    """Processa diretamente um item do DataPool na tela web simulada."""
    resultado = processar_lotes_playwright(
        [dados_lote],
        delay_passo=delay_passo,
        callback_log=callback_log,
        theme=theme,
        return_evidencias=True,
    )
    return resultado["evidencias"][0] if resultado["evidencias"] else None


def processar_datapool_playwright(
    delay_passo=0,
    callback_log=None,
    theme="dark",
    return_evidencias=False,
):
    """Processa em lote a planilha local, usado por demo e execucao isolada."""
    lotes = carregar_datapool_tratado()

    if not lotes:
        emitir_log("DataPool esta vazio ou a planilha nao foi localizada.", "warn", callback_log)
        return {"total": 0, "evidencias": []} if return_evidencias else 0

    return processar_lotes_playwright(
        lotes,
        delay_passo=delay_passo,
        callback_log=callback_log,
        theme=theme,
        return_evidencias=return_evidencias,
    )


def processar_lotes_playwright(
    lotes,
    delay_passo=0,
    callback_log=None,
    theme="dark",
    return_evidencias=False,
):
    url = obter_url_automacao()
    headless = os.getenv("PLAYWRIGHT_HEADLESS", "false").lower() == "true"
    base_referencia = carregar_base_referencia(ARQUIVO_INSPECAO)

    emitir_log(
        f"Iniciando execucao Playwright | Total de lotes: {len(lotes)} | Tema: {theme.upper()}",
        "info",
        callback_log,
    )

    with sync_playwright() as p:
        browser = iniciar_browser(p)
        context = browser.new_context(no_viewport=not headless)
        page = context.new_page()

        try:
            emitir_log(f"Acessando a URL do formulario: {url}", "info", callback_log)
            page.goto(url)

            login_page = LoginPagePlaywright(page, delay_passo=delay_passo)
            form_page = FormPagePlaywright(page, delay_passo=delay_passo)
            form_page.aplicar_tema(theme)

            emitir_log("Realizando autenticacao na plataforma...", "info", callback_log)
            login_page.fazer_login()

            processados = 0
            evidencias = []
            for idx, item in enumerate(lotes, start=1):
                msg_lote = (
                    f"[{idx:02d}/{len(lotes):02d}] Lote: {item['lote_id']} | "
                    f"Produto: {item['produto']} | Status: {item['status']}"
                )
                emitir_log(msg_lote, "info", callback_log)

                form_page.resetar_formulario()
                form_page.preencher_lote(item)
                sucesso = form_page.submeter_e_aguardar()
                resultado_item = analisar_dados_lote(item, base_referencia)
                form_page.registrar_analises(
                    resultado_item["analises"],
                    item,
                    resultado_item.get("linha_planilha"),
                )

                form_page.preparar_evidencia_visual()
                caminho_screenshot = montar_caminho_screenshot(item, "playwright")
                page.screenshot(path=str(caminho_screenshot), full_page=True)
                item["screenshot"] = str(caminho_screenshot)
                evidencias.append(
                    {
                        "lote_id": item.get("lote_id"),
                        "screenshot": str(caminho_screenshot),
                        "driver": "playwright",
                        "linha_planilha": resultado_item.get("linha_planilha"),
                        "analises": resultado_item["analises"],
                        "divergencias": resultado_item["divergencias"],
                        "avisos": resultado_item["avisos"],
                    }
                )

                if sucesso:
                    emitir_log(f"Lote {item['lote_id']} gravado com SUCESSO.", "success", callback_log)
                else:
                    emitir_log(f"Lote {item['lote_id']} submetido com ALERTAS.", "warn", callback_log)

                processados += 1

            emitir_log(
                f"Processamento finalizado com sucesso! Total: {processados} lotes",
                "success",
                callback_log,
            )
            if return_evidencias:
                return {"total": processados, "evidencias": evidencias}
            return processados

        except Exception as erro:
            emitir_log(f"Falha critica no Playwright: {erro}", "error", callback_log)
            raise
        finally:
            try:
                browser.close()
            except Exception:
                pass


if __name__ == "__main__":
    processar_datapool_playwright()
