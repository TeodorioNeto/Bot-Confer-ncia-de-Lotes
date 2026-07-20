"""
RN03 - Verificação de lote na base de referência (PDD seção 12).

"""

from pathlib import Path
import openpyxl


def carregar_base_referencia(
    caminho="dados_entrada/Inspecao_lotes_Formulario_Analise-Matriz-Priorização.xlsx",
    aba="Base_Referencia",
):
    """
    Carrega os lote_id cadastrados na aba de base de referência.

    Retorna um set com todos os lote_id oficialmente cadastrados.
    """
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(
            f"Base de referência não encontrada em: {caminho}"
        )

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        raise ValueError(f"Aba '{aba}' não encontrada no arquivo de referência")

    ws = wb[aba]
    linhas = ws.iter_rows(values_only=True)

    # Pula a linha de título (ex: "BASE DE REFERÊNCIA DE LOTES ...")
    next(linhas)
    cabecalho = next(linhas)
    idx_lote = cabecalho.index("lote_id")

    base = set()
    for linha in linhas:
        if idx_lote < len(linha) and linha[idx_lote]:
            base.add(linha[idx_lote])

    wb.close()
    return base


def verificar_lote_na_base(lote_id, base_referencia):
    """
    RN03: verifica se o lote_id existe na base de referência de lotes
    cadastrados.

    Args:
        lote_id: identificador do lote a verificar (ex: 'LG-2026-00101')
        base_referencia: set de lote_id cadastrados, geralmente vindo
                          de carregar_base_referencia()

    Returns:
        True se o lote_id existe na base, False caso contrário.

    Raises:
        ValueError: se lote_id estiver vazio.
    """
    if not lote_id:
        raise ValueError("lote_id é obrigatório (RN02/RN03)")

    return lote_id.strip() in base_referencia