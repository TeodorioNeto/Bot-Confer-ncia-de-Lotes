"""Analisa a planilha de lotes e preenche as abas de saida."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, PatternFill

from bot import processar_item
from src.base_referencia import carregar_base_referencia


ABA_INSPECAO = "Inspecao_14_06_2026"
ABA_ANALISE = "Formulario_Analise"
ABA_LOTES_AMBIGUOS = "lotes_ambiguos"
ABA_RESUMO_DIARIO = "Resumo_Diario"
COLUNAS = [
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
]

PREENCHIMENTO_DIVERGENCIA = PatternFill("solid", fgColor="FCE8E6")
PREENCHIMENTO_AVISO = PatternFill("solid", fgColor="FFF4CC")
CABECALHO_FILL = PatternFill("solid", fgColor="1F4E78")


class ItemPlanilha:
    """Adaptador de uma linha da planilha para a interface do DataPoolEntry."""

    def __init__(self, valores):
        self.valores = valores

    def get_value(self, chave):
        return self.valores.get(chave)


def analisar_e_preencher_formulario(caminho_entrada, caminho_saida):
    """Executa as regras e salva uma cópia com o formulário preenchido."""
    caminho_entrada = Path(caminho_entrada)
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    base_referencia = carregar_base_referencia(caminho_entrada)
    workbook = openpyxl.load_workbook(caminho_entrada)
    try:
        if ABA_INSPECAO not in workbook.sheetnames:
            raise ValueError(f"Aba '{ABA_INSPECAO}' não encontrada")
        if ABA_ANALISE not in workbook.sheetnames:
            raise ValueError(f"Aba '{ABA_ANALISE}' não encontrada")

        resultados = _analisar_registros(workbook[ABA_INSPECAO], base_referencia)
        resumo = _preencher_aba(workbook[ABA_ANALISE], resultados)
        resumo["lotes_ambiguos"] = _preencher_lotes_ambiguos(workbook, resultados)
        _preencher_resumo_diario(workbook, resumo)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(caminho_saida)
    finally:
        workbook.close()

    return caminho_saida, resultados, resumo


def _analisar_registros(ws, base_referencia):
    linha_cabecalho = _encontrar_cabecalho(ws)
    resultados = []

    for numero_linha in range(linha_cabecalho + 1, ws.max_row + 1):
        valores_linha = [ws.cell(numero_linha, coluna).value for coluna in range(1, 9)]
        preenchidos = sum(
            valor is not None and bool(str(valor).strip())
            for valor in valores_linha
        )
        if preenchidos == 0:
            break
        if preenchidos < 4:
            continue

        valores = dict(zip(COLUNAS, valores_linha))
        resultado = processar_item(ItemPlanilha(valores), base_referencia)
        resultado["registro"] = len(resultados) + 1
        resultado["linha_planilha"] = numero_linha
        resultado["status"] = valores.get("status")
        resultados.append(resultado)

    return resultados


def _encontrar_cabecalho(ws):
    for numero_linha in range(1, ws.max_row + 1):
        valores = [ws.cell(numero_linha, coluna).value for coluna in range(1, 9)]
        if valores == COLUNAS:
            return numero_linha
    raise ValueError("Cabeçalho da inspeção não encontrado")


def _preencher_aba(ws, resultados):
    primeira_linha = 4
    ultima_linha = primeira_linha + len(resultados) - 1

    for numero_linha in range(primeira_linha, ws.max_row + 1):
        if numero_linha > 28:
            break
        for coluna in range(2, 7):
            ws.cell(numero_linha, coluna).value = None

    for resultado in resultados:
        numero_linha = primeira_linha + resultado["registro"] - 1
        analises = resultado["analises"]
        if not analises:
            continue

        ws.cell(numero_linha, 2).value = resultado["lote_id"] or "(vazio)"
        ws.cell(numero_linha, 3).value = " | ".join(
            analise["problema"] for analise in analises
        )
        ws.cell(numero_linha, 4).value = " / ".join(
            dict.fromkeys(analise["regra"] for analise in analises)
        )
        ws.cell(numero_linha, 5).value = " | ".join(
            analise["acao"] for analise in analises
        )
        categoria = "divergência" if resultado["divergencias"] else "aviso"
        ws.cell(numero_linha, 6).value = f"Sim ({categoria})"

        preenchimento = (
            PREENCHIMENTO_DIVERGENCIA
            if resultado["divergencias"]
            else PREENCHIMENTO_AVISO
        )
        for coluna in range(2, 7):
            celula = ws.cell(numero_linha, coluna)
            celula.fill = preenchimento
            celula.alignment = Alignment(vertical="top", wrap_text=True)

    ws["D31"] = f"=COUNTA(A{primeira_linha}:A{ultima_linha})"
    ws["D32"] = f"=COUNTBLANK(C{primeira_linha}:C{ultima_linha})"
    ws["D33"] = f'=COUNTIF(F{primeira_linha}:F{ultima_linha},"*divergência*")'
    ws["D34"] = f'=COUNTIF(C{primeira_linha}:C{ultima_linha},"*APROVADO PARCIAL*")'
    ws["D35"] = f'=COUNTIF(D{primeira_linha}:D{ultima_linha},"*RN05*")'
    ws["D36"] = "+".join(
        f'COUNTIF(D{primeira_linha}:D{ultima_linha},"*{regra}*")'
        for regra in ("RN02", "RN03", "RN04", "RN06", "RN07")
    )
    ws["D36"] = f"={ws['D36'].value}"
    ws["D37"] = "=IF(D31=0,0,D33/D31)"
    ws["D37"].number_format = "0.00%"
    ws["D38"] = "Automatizar com revisão humana dos casos ambíguos"

    for linha in range(31, 39):
        ws.cell(linha, 4).alignment = Alignment(vertical="center", wrap_text=True)

    registros_com_ocorrencia = sum(bool(r["analises"]) for r in resultados)
    registros_com_divergencia = sum(bool(r["divergencias"]) for r in resultados)
    total_divergencias = sum(len(r["divergencias"]) for r in resultados)
    normalizacoes = sum(
        any(a["regra"] == "RN05" for a in r["analises"]) for r in resultados
    )
    return {
        "total_registros": len(resultados),
        "registros_validos": len(resultados) - registros_com_ocorrencia,
        "registros_com_ocorrencia": registros_com_ocorrencia,
        "registros_com_divergencia": registros_com_divergencia,
        "normalizacoes": normalizacoes,
        "total_divergencias": total_divergencias,
    }


def _preencher_lotes_ambiguos(workbook, resultados):
    ws = _obter_ou_criar_aba(workbook, ABA_LOTES_AMBIGUOS)
    _limpar_aba(ws)
    cabecalho = [
        "registro",
        "linha_planilha",
        "lote_id",
        "status",
        "regra",
        "problema",
        "acao_recomendada",
    ]
    _escrever_cabecalho(ws, cabecalho)

    total = 0
    for resultado in resultados:
        for analise in resultado["analises"]:
            if analise["regra"] != "RN06":
                continue

            total += 1
            ws.append(
                [
                    resultado["registro"],
                    resultado["linha_planilha"],
                    resultado["lote_id"] or "(vazio)",
                    resultado.get("status") or "",
                    analise["regra"],
                    analise["problema"],
                    analise["acao"],
                ]
            )

    _ajustar_colunas(ws)
    return total


def _preencher_resumo_diario(workbook, resumo):
    ws = _obter_ou_criar_aba(workbook, ABA_RESUMO_DIARIO)
    _limpar_aba(ws)
    _escrever_cabecalho(ws, ["indicador", "valor"])

    linhas = [
        ("total_registros", resumo["total_registros"]),
        ("registros_validos", resumo["registros_validos"]),
        ("registros_com_ocorrencia", resumo["registros_com_ocorrencia"]),
        ("registros_com_divergencia", resumo["registros_com_divergencia"]),
        ("lotes_ambiguos", resumo["lotes_ambiguos"]),
        ("normalizacoes", resumo["normalizacoes"]),
        ("total_divergencias", resumo["total_divergencias"]),
    ]
    for indicador, valor in linhas:
        ws.append([indicador, valor])

    if resumo["total_registros"]:
        taxa = resumo["registros_com_divergencia"] / resumo["total_registros"]
    else:
        taxa = 0
    ws.append(["taxa_divergencia", taxa])
    ws["B9"].number_format = "0.00%"

    _ajustar_colunas(ws)


def _obter_ou_criar_aba(workbook, nome):
    if nome in workbook.sheetnames:
        return workbook[nome]
    return workbook.create_sheet(nome)


def _limpar_aba(ws):
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def _escrever_cabecalho(ws, valores):
    ws.append(valores)
    for celula in ws[1]:
        celula.fill = CABECALHO_FILL
        celula.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_colunas(ws):
    for coluna in ws.columns:
        largura = max(
            (len(str(celula.value)) for celula in coluna if celula.value is not None),
            default=10,
        )
        ws.column_dimensions[coluna[0].column_letter].width = min(largura + 2, 60)
