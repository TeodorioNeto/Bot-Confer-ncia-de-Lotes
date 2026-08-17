import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from bot import processar_item
from config import ARQUIVO_INSPECAO
from src.base_referencia import carregar_base_referencia
from src.validacao import COLUNAS_ESTRUTURA
from src.web_evidencias import obter_url_automacao


load_dotenv()


def ambiente_container():
    """Indica se a automacao esta rodando dentro de container."""
    return os.getenv("ENVIRONMENT", "local").lower() != "local"


def iniciar_browser(playwright):
    """Inicializa Chromium com flags adequadas para execucao local ou container."""
    em_container = ambiente_container()
    headless = em_container or os.getenv("PLAYWRIGHT_HEADLESS", "false").lower() == "true"
    argumentos = []

    if em_container:
        argumentos.extend(
            ["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"]
        )
    else:
        argumentos.append("--start-maximized")

    opcoes = {"headless": headless, "args": argumentos}
    if not em_container:
        opcoes["channel"] = os.getenv("PLAYWRIGHT_CHANNEL", "msedge")

    return playwright.chromium.launch(**opcoes)


def montar_dados_lote(item):
    """Converte um item do DataPool para o formato usado pela automacao web."""
    dados = {campo: item.get_value(campo) for campo in COLUNAS_ESTRUTURA}
    dados["lote"] = dados.get("lote_id")
    if hasattr(item, "get_value"):
        dados["linha_planilha"] = item.get_value("linha_planilha")
    return dados


def analisar_dados_lote(dados_lote, base_referencia=None):
    """Executa bot.py para gerar as mesmas analises usadas na planilha."""
    base_referencia = base_referencia or carregar_base_referencia(ARQUIVO_INSPECAO)
    resultado = processar_item(ItemPlanilhaWeb(dados_lote), base_referencia)
    resultado["linha_planilha"] = dados_lote.get("linha_planilha")
    resultado["status"] = dados_lote.get("status")
    return resultado


def processar_item_web(
    item,
    driver=None,
    delay_passo=None,
    callback_log=None,
    theme="dark",
):
    """Executa a automacao web usando diretamente o item atual do DataPool."""
    driver = (driver or os.getenv("WEB_AUTOMATION_DRIVER", "playwright")).lower()
    dados_lote = montar_dados_lote(item)

    if driver == "selenium":
        from src.web_automation_selenium import processar_item_selenium

        kwargs = {"callback_log": callback_log, "theme": theme}
        if delay_passo is not None:
            kwargs["delay_passo"] = delay_passo
        return processar_item_selenium(dados_lote, **kwargs)

    if driver == "playwright":
        from src.web_automation_playwright import processar_item_playwright

        kwargs = {"callback_log": callback_log, "theme": theme}
        if delay_passo is not None:
            kwargs["delay_passo"] = delay_passo
        return processar_item_playwright(dados_lote, **kwargs)

    raise ValueError(
        "WEB_AUTOMATION_DRIVER deve ser 'playwright' ou 'selenium'. "
        f"Valor recebido: {driver}"
    )


def processar_datapool(
    driver=None,
    delay_passo=None,
    callback_log=None,
    theme="dark",
    return_evidencias=False,
):
    """Executa a automacao web no fluxo atual baseado em DataPool/planilha."""
    driver = (driver or os.getenv("WEB_AUTOMATION_DRIVER", "playwright")).lower()

    if driver == "selenium":
        from src.web_automation_selenium import processar_datapool_selenium

        kwargs = {"callback_log": callback_log, "theme": theme}
        if delay_passo is not None:
            kwargs["delay_passo"] = delay_passo
        if return_evidencias:
            kwargs["return_evidencias"] = True
        return processar_datapool_selenium(**kwargs)

    if driver == "playwright":
        from src.web_automation_playwright import processar_datapool_playwright

        kwargs = {"callback_log": callback_log, "theme": theme}
        if delay_passo is not None:
            kwargs["delay_passo"] = delay_passo
        if return_evidencias:
            kwargs["return_evidencias"] = True
        return processar_datapool_playwright(**kwargs)

    raise ValueError(
        "WEB_AUTOMATION_DRIVER deve ser 'playwright' ou 'selenium'. "
        f"Valor recebido: {driver}"
    )


def _dados_lote_demo():
    return {
        "lote_id": "LOTE-2026-0001",
        "produto": "Placa Mae V1",
        "linha": "A",
        "turno": "MANHA",
        "status": "APROVADO",
        "responsavel": "Usuario Teste",
        "data": "2026-07-27",
        "observacao": "",
    }


def carregar_primeiro_lote_da_planilha(caminho_planilha=None):
    """
    Carrega um lote da planilha de entrada usada pelo Dispatcher/BotCity.

    Essa funcao existe para que a execucao isolada de `python -m src.web_automation`
    use dados reais da planilha local, sem depender de um item do DataPool.
    """
    caminho = Path(caminho_planilha or ARQUIVO_INSPECAO)
    if not caminho.exists():
        return _dados_lote_demo()

    workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        for ws in workbook.worksheets:
            linha_cabecalho = _encontrar_linha_cabecalho(ws)
            if linha_cabecalho is None:
                continue

            for linha in ws.iter_rows(
                min_row=linha_cabecalho + 1,
                max_col=len(COLUNAS_ESTRUTURA),
                values_only=True,
            ):
                if linha is None or all(valor is None for valor in linha):
                    break

                preenchidos = sum(
                    valor is not None and bool(str(valor).strip()) for valor in linha
                )
                if preenchidos >= 4:
                    return dict(zip(COLUNAS_ESTRUTURA, linha))
    finally:
        workbook.close()

    return _dados_lote_demo()


def carregar_primeiro_resultado_da_planilha(caminho_planilha=None):
    """Carrega da planilha o primeiro lote com ocorrencia; se nao houver, usa o primeiro lote."""
    return carregar_resultado_planilha_para_web(caminho_planilha)


def carregar_resultado_planilha_para_web(caminho_planilha=None):
    """Carrega a planilha inteira e prepara todas as ocorrencias para o simulador web."""
    caminho = Path(caminho_planilha or ARQUIVO_INSPECAO)
    if not caminho.exists():
        return {"dados_lote": _dados_lote_demo(), "analises": [], "linha_planilha": ""}

    base_referencia = carregar_base_referencia(caminho)
    primeiro_lote = None
    primeiro_lote_com_ocorrencia = None
    primeira_linha_com_ocorrencia = ""
    analises_web = []

    workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        for ws in workbook.worksheets:
            linha_cabecalho = _encontrar_linha_cabecalho(ws)
            if linha_cabecalho is None:
                continue

            for numero_linha in range(linha_cabecalho + 1, ws.max_row + 1):
                valores_linha = [
                    ws.cell(numero_linha, coluna).value
                    for coluna in range(1, len(COLUNAS_ESTRUTURA) + 1)
                ]
                if valores_linha is None or all(valor is None for valor in valores_linha):
                    break

                preenchidos = sum(
                    valor is not None and bool(str(valor).strip())
                    for valor in valores_linha
                )
                if preenchidos < 4:
                    continue

                dados_lote = dict(zip(COLUNAS_ESTRUTURA, valores_linha))
                resultado = processar_item(ItemPlanilhaWeb(dados_lote), base_referencia)

                if primeiro_lote is None:
                    primeiro_lote = dados_lote
                if resultado["analises"]:
                    if primeiro_lote_com_ocorrencia is None:
                        primeiro_lote_com_ocorrencia = dados_lote
                        primeira_linha_com_ocorrencia = numero_linha

                    for analise in resultado["analises"]:
                        analise_web = dict(analise)
                        analise_web["linha_planilha"] = numero_linha
                        analise_web["lote_id"] = resultado["lote_id"] or "(vazio)"
                        analises_web.append(analise_web)
    finally:
        workbook.close()

    return {
        "dados_lote": primeiro_lote_com_ocorrencia or primeiro_lote or _dados_lote_demo(),
        "analises": analises_web,
        "linha_planilha": primeira_linha_com_ocorrencia,
    }


def _encontrar_linha_cabecalho(ws):
    for numero_linha in range(1, ws.max_row + 1):
        valores = [
            ws.cell(numero_linha, coluna).value
            for coluna in range(1, len(COLUNAS_ESTRUTURA) + 1)
        ]
        if valores == COLUNAS_ESTRUTURA:
            return numero_linha
    return None


class ItemPlanilhaWeb:
    def __init__(self, valores):
        self.valores = valores

    def get_value(self, chave):
        return self.valores.get(chave)


if __name__ == "__main__":
    processar_datapool()
