import tempfile
import unittest
from pathlib import Path

import pandas as pd

from src.validacao import carregar_planilha, valida_campos_obrigatorios, valida_estrutura


COLUNAS_VALIDAS = {
    "lote_id": ["L001", "L002"],
    "produto": ["TV", "Monitor"],
    "linha": ["A", "B"],
    "turno": ["Manha", "Tarde"],
    "status": ["OK", "NOK"],
    "responsavel": ["Ana", "Bruno"],
    "data": ["2026-07-14", "2026-07-14"],
    "observacao": ["", "Falha encontrada"],
}


class TestValidacaoRN01RN02(unittest.TestCase):
    def criar_planilha(self, dados):
        temp_dir = tempfile.TemporaryDirectory()
        caminho = Path(temp_dir.name) / "lotes.xlsx"
        pd.DataFrame(dados).to_excel(caminho, index=False)
        self.addCleanup(temp_dir.cleanup)
        return caminho

    def test_valida_estrutura_quando_colunas_obrigatorias_existem(self):
        caminho = self.criar_planilha(COLUNAS_VALIDAS)

        self.assertTrue(valida_estrutura(caminho))

    def test_reprova_estrutura_quando_coluna_obrigatoria_falta(self):
        dados = dict(COLUNAS_VALIDAS)
        dados.pop("status")
        caminho = self.criar_planilha(dados)

        self.assertFalse(valida_estrutura(caminho))

    def test_reprova_estrutura_quando_existe_coluna_extra(self):
        dados = dict(COLUNAS_VALIDAS)
        dados["coluna_extra"] = ["x", "y"]
        caminho = self.criar_planilha(dados)

        self.assertFalse(valida_estrutura(caminho))

    def test_reprova_estrutura_quando_colunas_estao_fora_de_ordem(self):
        dados = {
            "produto": ["TV"],
            "lote_id": ["L001"],
            "linha": ["A"],
            "turno": ["Manha"],
            "status": ["OK"],
            "responsavel": ["Ana"],
            "data": ["2026-07-14"],
            "observacao": [""],
        }
        caminho = self.criar_planilha(dados)

        self.assertFalse(valida_estrutura(caminho))

    def test_valida_campos_obrigatorios_quando_todos_estao_preenchidos(self):
        caminho = self.criar_planilha(COLUNAS_VALIDAS)

        self.assertTrue(valida_campos_obrigatorios(caminho))

    def test_reutiliza_dataframe_carregado_nas_validacoes(self):
        caminho = self.criar_planilha(COLUNAS_VALIDAS)
        df = carregar_planilha(caminho)

        self.assertTrue(valida_estrutura(df=df))
        self.assertTrue(valida_campos_obrigatorios(df=df))

    def test_reprova_campos_obrigatorios_quando_existe_vazio(self):
        dados = dict(COLUNAS_VALIDAS)
        dados["responsavel"] = ["Ana", ""]
        caminho = self.criar_planilha(dados)

        self.assertFalse(valida_campos_obrigatorios(caminho))

    def test_reprova_arquivo_que_nao_e_xlsx(self):
        with self.assertRaises(ValueError):
            valida_estrutura("lotes.csv")

    def test_valida_estrutura_com_linhas_descritivas_antes_do_cabecalho(self):
        temp_dir = tempfile.TemporaryDirectory()
        caminho = Path(temp_dir.name) / "lotes.xlsx"
        self.addCleanup(temp_dir.cleanup)

        linhas = [
            ["PLANILHA DE INSPECAO DE LOTES"],
            ["Arquivo gerado pelo sistema"],
            list(COLUNAS_VALIDAS.keys()),
            ["L001", "TV", "A", "Manha", "OK", "Ana", "2026-07-14", ""],
        ]
        pd.DataFrame(linhas).to_excel(caminho, index=False, header=False)

        self.assertTrue(valida_estrutura(caminho))

    def test_ignora_rodape_apos_linhas_de_registro(self):
        temp_dir = tempfile.TemporaryDirectory()
        caminho = Path(temp_dir.name) / "lotes.xlsx"
        self.addCleanup(temp_dir.cleanup)

        linhas = [
            list(COLUNAS_VALIDAS.keys()),
            ["L001", "TV", "A", "Manha", "OK", "Ana", "2026-07-14", ""],
            ["Total de registros: 1", "", "", "", "Resumo final", "", "", ""],
        ]
        pd.DataFrame(linhas).to_excel(caminho, index=False, header=False)

        self.assertTrue(valida_campos_obrigatorios(caminho))

if __name__ == "__main__":
    unittest.main()
import pytest
import openpyxl
from bot import processar_item
from src.base_referencia import verificar_lote_na_base, carregar_base_referencia
from src.validacao import (
    normalizar_status,
    status_ambiguo,
    valida_observacao_reprovado,
    valida_status,
)
from src.relatorio import gerar_relatorio_divergencias

@pytest.fixture
def base_exemplo():
    return {"LG-2026-00101", "LG-2026-00102", "LG-2026-00104", "LG-2026-00105"}


class ItemTeste:
    def __init__(self, valores):
        self.valores = valores

    def get_value(self, chave):
        return self.valores.get(chave)


def criar_item(status, observacao="Inspecao conferida"):
    return ItemTeste(
        {
            "lote_id": "LG-2026-00101",
            "produto": "TV",
            "linha": "A",
            "turno": "MANHA",
            "status": status,
            "responsavel": "Ana",
            "data": "14/06/2026",
            "observacao": observacao,
        }
    )


def test_lote_existente_na_base(base_exemplo):
    assert verificar_lote_na_base("LG-2026-00101", base_exemplo) is True


def test_lote_inexistente_na_base(base_exemplo):
    # Caso do gabarito: LG-2026-00103 não existe na Base_Referencia (RN03)
    assert verificar_lote_na_base("LG-2026-00103", base_exemplo) is False


def test_lote_id_vazio_gera_erro(base_exemplo):
    with pytest.raises(ValueError):
        verificar_lote_na_base("", base_exemplo)


def test_carregar_base_referencia_arquivo_inexistente(tmp_path):
    with pytest.raises(FileNotFoundError):
        carregar_base_referencia(str(tmp_path / "nao_existe.xlsx"))


def test_carregar_base_referencia_com_arquivo_temporario(tmp_path):
    """
    Testa carregar_base_referencia() com um xlsx fictício criado no teste,
    sem depender da planilha real (que não é versionada no Git).
    """
    

    caminho = tmp_path / "base_teste.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base_Referencia"
    ws.append(["BASE DE REFERÊNCIA DE LOTES (título)"])  # linha de título, ignorada
    ws.append(["lote_id", "codigo_produto", "descricao_produto", "status_cadastro"])
    ws.append(["LG-2026-00101", "COD01", "Setupbox modelo A", "ativo"])
    ws.append(["LG-2026-00102", "COD02", "Setupbox modelo B", "ativo"])
    wb.save(caminho)

    base = carregar_base_referencia(caminho=str(caminho), aba="Base_Referencia")

    assert "LG-2026-00101" in base
    assert "LG-2026-00103" not in base


def test_normaliza_ok_para_aprovado():
    assert normalizar_status("OK") == "APROVADO"


def test_normaliza_nok_para_reprovado():
    assert normalizar_status("NOK") == "REPROVADO"


def test_normaliza_ignora_case_e_espacos():
    assert normalizar_status("  ok  ") == "APROVADO"


def test_normaliza_mantem_status_ja_oficial():
    assert normalizar_status("APROVADO") == "APROVADO"
    assert normalizar_status("PENDENTE") == "PENDENTE"


def test_valida_status_aceita_valores_oficiais():
    assert valida_status("APROVADO") is True
    assert valida_status("REPROVADO") is True
    assert valida_status("PENDENTE") is True


def test_valida_status_aceita_sinonimos_normalizados():
    assert valida_status("OK") is True
    assert valida_status("NOK") is True


def test_valida_status_rejeita_valor_nao_reconhecido():
    # Casos reais do gabarito: "REPROV." e "APROVADO PARCIAL" (RN06)
    assert valida_status("REPROV.") is False
    assert valida_status("APROVADO PARCIAL") is False


def test_rn06_identifica_status_ambiguo():
    assert status_ambiguo("REPROV.") is True
    assert status_ambiguo("APROVADO PARCIAL") is True


def test_rn06_nao_marca_status_oficial_ou_normalizavel():
    assert status_ambiguo("APROVADO") is False
    assert status_ambiguo("NOK") is False
    assert status_ambiguo("") is False


def test_performer_encaminha_status_ambiguo_para_revisao(base_exemplo):
    resultado = processar_item(criar_item("APROVADO PARCIAL"), base_exemplo)
    regras = [analise["regra"] for analise in resultado["analises"]]

    assert regras == ["RN04", "RN06"]
    assert resultado["analises"][1]["acao"] == "Encaminhar o registro para revisão humana"


def test_performer_normaliza_nok_e_exige_observacao(base_exemplo):
    resultado = processar_item(criar_item("NOK", observacao=""), base_exemplo)

    assert any(analise["regra"] == "RN05" for analise in resultado["analises"])
    assert any(analise["regra"] == "RN07" for analise in resultado["analises"])
    assert not any(analise["regra"] == "RN06" for analise in resultado["analises"])


def test_valida_status_vazio_gera_erro():
    with pytest.raises(ValueError):
        valida_status("")

    with pytest.raises(ValueError):
        valida_status(None)


def test_gera_relatorio_com_divergencias(tmp_path):
    divergencias = [
        {"linha": 6, "lote_id": "LG-2026-00103", "regra": "RN03", "problema": "lote_id não existe na base de referência"},
        {"linha": 27, "lote_id": None, "regra": "RN02", "problema": "lote_id vazio"},
    ]
    caminho_saida = tmp_path / "relatorio_teste.xlsx"

    resultado = gerar_relatorio_divergencias(divergencias, caminho_saida=caminho_saida)

    assert resultado.exists()

    wb = openpyxl.load_workbook(resultado)
    ws = wb.active
    assert ws["A1"].value == "Linha"
    assert ws["B1"].value == "Lote ID"
    assert ws["A2"].value == 6
    assert ws["B2"].value == "LG-2026-00103"
    assert ws["A3"].value == 27


def test_gera_relatorio_vazio_sem_divergencias(tmp_path):
    caminho_saida = tmp_path / "relatorio_vazio.xlsx"

    resultado = gerar_relatorio_divergencias([], caminho_saida=caminho_saida)

    assert resultado.exists()
    wb = openpyxl.load_workbook(resultado)
    ws = wb.active
    assert ws["A1"].value == "Linha"
    assert ws.max_row == 1  # só o cabeçalho, nenhuma divergência


def test_gera_relatorio_cria_pasta_data_output(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)

    resultado = gerar_relatorio_divergencias([])

    assert resultado.exists()
    assert "data" in str(resultado)
    assert "output" in str(resultado)


class TestValidacaoRN07(unittest.TestCase):
    def criar_planilha(self, linhas, cabecalho=None):
        temp_dir = tempfile.TemporaryDirectory()
        caminho = Path(temp_dir.name) / "lotes.xlsx"
        self.addCleanup(temp_dir.cleanup)

        cabecalho = cabecalho or list(COLUNAS_VALIDAS.keys())
        pd.DataFrame(linhas, columns=cabecalho).to_excel(caminho, index=False)
        return caminho

    def linha_valida(self, status, observacao):
        return ["LG-2026-00101", "TV", "A", "Manha", status, "Ana", "2026-07-14", observacao]

    def test_valida_reprovado_com_observacao_preenchida(self):
        caminho = self.criar_planilha([
            self.linha_valida("REPROVADO", "Defeito na tela")
        ])

        self.assertTrue(valida_observacao_reprovado(caminho))

    def test_reprova_reprovado_sem_observacao(self):
        caminho = self.criar_planilha([
            self.linha_valida("REPROVADO", "")
        ])

        self.assertFalse(valida_observacao_reprovado(caminho))

    def test_reprova_nok_sem_observacao(self):
        caminho = self.criar_planilha([
            self.linha_valida("NOK", "")
        ])

        self.assertFalse(valida_observacao_reprovado(caminho))

    def test_nao_exige_observacao_para_aprovado(self):
        caminho = self.criar_planilha([
            self.linha_valida("APROVADO", "")
        ])

        self.assertTrue(valida_observacao_reprovado(caminho))

    def test_reprova_quando_coluna_observacao_falta(self):
        caminho = self.criar_planilha(
            [["LG-2026-00105", "TV", "A", "Manha", "REPROVADO", "Ana", "2026-07-14"]],
            cabecalho=[coluna for coluna in COLUNAS_VALIDAS if coluna != "observacao"],
        )

        self.assertFalse(valida_observacao_reprovado(caminho))
