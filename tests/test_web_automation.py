from unittest.mock import Mock, patch
import src.web_automation_playwright
import src.web_automation_selenium
import openpyxl
import pytest
from src import web_automation, web_evidencias
from src.web_evidencias import montar_caminho_screenshot


class ItemWebTeste:
    def __init__(self, valores):
        self.valores = valores

    def get_value(self, chave):
        return self.valores.get(chave)


def test_monta_dados_lote_a_partir_do_item_datapool():
    item = ItemWebTeste(
        {
            "lote_id": "LG-2026-00101",
            "produto": "TV",
            "linha": "A",
            "turno": "MANHA",
            "status": "APROVADO",
            "responsavel": "Ana",
            "data": "14/06/2026",
            "observacao": "",
        }
    )

    dados = web_automation.montar_dados_lote(item)

    assert dados["lote_id"] == "LG-2026-00101"
    assert dados["produto"] == "TV"
    assert dados["status"] == "APROVADO"
    assert "observacao" in dados


def test_carrega_primeiro_lote_da_planilha_usada_pelo_botcity(tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["PLANILHA DE INSPECAO"])
    ws.append(["Arquivo", "Sistema", "Registros"])
    ws.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
        ]
    )
    ws.append(["LG-2026-00101", "TV", "A", "MANHA", "APROVADO", "Ana", "14/06/2026", ""])
    wb.save(caminho)

    dados = web_automation.carregar_primeiro_lote_da_planilha(caminho)

    assert dados["lote_id"] == "LG-2026-00101"
    assert dados["produto"] == "TV"
    assert dados["status"] == "APROVADO"


def test_carregar_datapool_tratado_usa_cabecalho_real_da_planilha(monkeypatch, tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["PLANILHA DE INSPECAO"])
    ws.append(["Arquivo", "Sistema", "Registros"])
    ws.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
        ]
    )
    ws.append(["LG-2026-00101", "TV55-4K-B", "L1", "A", "APROVADO", "Ana", "14/06/2026", ""])
    wb.save(caminho)

    monkeypatch.setattr(src.web_automation_playwright, "ARQUIVO_INSPECAO", caminho)

    lotes = src.web_automation_playwright.carregar_datapool_tratado()

    assert lotes[0]["lote_id"] == "LG-2026-00101"
    assert lotes[0]["produto"] == "TV55-4K-B"
    assert lotes[0]["status"] == "APROVADO"


def test_carrega_primeiro_resultado_com_ocorrencia_para_formulario_analise(tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["PLANILHA DE INSPECAO"])
    ws.append(["Arquivo", "Sistema", "Registros"])
    ws.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
        ]
    )
    ws.append(["LG-2026-00101", "TV", "A", "MANHA", "APROVADO", "Ana", "14/06/2026", ""])
    ws.append(["LG-2026-00102", "TV", "A", "MANHA", "NOK", "Bia", "14/06/2026", ""])

    base = wb.create_sheet("Base_Referencia")
    base.append(["BASE DE REFERENCIA"])
    base.append(["lote_id", "codigo_produto", "descricao_produto", "status_cadastro"])
    base.append(["LG-2026-00101", "P001", "TV", "ativo"])
    base.append(["LG-2026-00102", "P002", "TV", "ativo"])
    wb.save(caminho)

    resultado = web_automation.carregar_primeiro_resultado_da_planilha(caminho)

    assert resultado["dados_lote"]["lote_id"] == "LG-2026-00102"
    assert resultado["linha_planilha"] == 5
    assert any(analise["regra"] == "RN07" for analise in resultado["analises"])


def test_carrega_todas_ocorrencias_da_planilha_para_formulario_analise(tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["PLANILHA DE INSPECAO"])
    ws.append(["Arquivo", "Sistema", "Registros"])
    ws.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
        ]
    )
    ws.append(["LG-2026-00101", "TV", "A", "MANHA", "APROVADO", "Ana", "14/06/2026", ""])
    ws.append(["LG-2026-00102", "TV", "A", "MANHA", "NOK", "Bia", "14/06/2026", ""])
    ws.append(["LG-2026-00103", "TV", "A", "MANHA", "APROVADO PARCIAL", "Caio", "14/06/2026", ""])

    base = wb.create_sheet("Base_Referencia")
    base.append(["BASE DE REFERENCIA"])
    base.append(["lote_id", "codigo_produto", "descricao_produto", "status_cadastro"])
    base.append(["LG-2026-00101", "P001", "TV", "ativo"])
    base.append(["LG-2026-00102", "P002", "TV", "ativo"])
    base.append(["LG-2026-00103", "P003", "TV", "ativo"])
    wb.save(caminho)

    resultado = web_automation.carregar_resultado_planilha_para_web(caminho)
    regras = [analise["regra"] for analise in resultado["analises"]]
    linhas = [analise["linha_planilha"] for analise in resultado["analises"]]

    assert resultado["dados_lote"]["lote_id"] == "LG-2026-00102"
    assert "RN07" in regras
    assert "RN04" in regras
    assert "RN06" in regras
    assert 5 in linhas
    assert 6 in linhas


def test_processar_datapool_usa_playwright_por_padrao(monkeypatch):
    monkeypatch.delenv("WEB_AUTOMATION_DRIVER", raising=False)
    processar_playwright = Mock(return_value=2)

    with patch(
        "src.web_automation_playwright.processar_datapool_playwright",
        processar_playwright,
    ):
        resultado = web_automation.processar_datapool()

    assert resultado == 2
    processar_playwright.assert_called_once_with(callback_log=None, theme="dark")


def test_processar_datapool_usa_selenium_quando_configurado(monkeypatch):
    monkeypatch.setenv("WEB_AUTOMATION_DRIVER", "selenium")
    processar_selenium = Mock(return_value=3)

    with patch(
        "src.web_automation_selenium.processar_datapool_selenium",
        processar_selenium,
    ):
        resultado = web_automation.processar_datapool()

    assert resultado == 3
    processar_selenium.assert_called_once_with(callback_log=None, theme="dark")


def test_criar_driver_selenium_usa_executavel_local(monkeypatch, tmp_path):
    caminho_driver = tmp_path / "msedgedriver.exe"
    caminho_driver.touch()
    monkeypatch.setenv("EDGE_DRIVER_PATH", str(caminho_driver))
    driver = Mock()

    with (
        patch("src.web_automation_selenium.EdgeService") as service_class,
        patch("src.web_automation_selenium.webdriver.Edge", return_value=driver),
        patch(
            "src.web_automation_selenium.EdgeChromiumDriverManager.install"
        ) as instalar_automaticamente,
    ):
        resultado = src.web_automation_selenium.criar_driver()

    service_class.assert_called_once_with(str(caminho_driver.resolve()))
    instalar_automaticamente.assert_not_called()
    driver.maximize_window.assert_called_once_with()
    assert resultado is driver


def test_criar_driver_selenium_falha_se_executavel_local_nao_existe(
    monkeypatch, tmp_path
):
    caminho_driver = tmp_path / "msedgedriver.exe"
    monkeypatch.setenv("EDGE_DRIVER_PATH", str(caminho_driver))

    with pytest.raises(FileNotFoundError, match="EdgeDriver nao encontrado"):
        src.web_automation_selenium.criar_driver()


def test_processar_datapool_rejeita_driver_desconhecido():
    with pytest.raises(ValueError):
        web_automation.processar_datapool(driver="desconhecido")


def test_processar_datapool_repassa_opcoes_da_nova_versao(monkeypatch):
    monkeypatch.setenv("WEB_AUTOMATION_DRIVER", "selenium")
    processar_selenium = Mock(return_value=4)
    callback_log = Mock()

    with patch(
        "src.web_automation_selenium.processar_datapool_selenium",
        processar_selenium,
    ):
        resultado = web_automation.processar_datapool(
            delay_passo=0.4,
            callback_log=callback_log,
            theme="light",
        )

    assert resultado == 4
    processar_selenium.assert_called_once_with(
        delay_passo=0.4,
        callback_log=callback_log,
        theme="light",
    )


def test_processar_datapool_repassa_retorno_de_evidencias(monkeypatch):
    monkeypatch.setenv("WEB_AUTOMATION_DRIVER", "playwright")
    retorno = {
        "total": 1,
        "evidencias": [
            {
                "lote_id": "LG-2026-00101",
                "screenshot": "logs/screenshots/playwright/evidencia.png",
                "driver": "playwright",
            }
        ],
    }
    processar_playwright = Mock(return_value=retorno)

    with patch(
        "src.web_automation_playwright.processar_datapool_playwright",
        processar_playwright,
    ):
        resultado = web_automation.processar_datapool(return_evidencias=True)

    assert resultado == retorno
    processar_playwright.assert_called_once_with(
        callback_log=None,
        theme="dark",
        return_evidencias=True,
    )


def test_processar_item_web_usa_item_datapool_diretamente(monkeypatch):
    monkeypatch.setenv("WEB_AUTOMATION_DRIVER", "playwright")
    item = ItemWebTeste(
        {
            "lote_id": "LG-2026-00101",
            "produto": "TV",
            "linha": "A",
            "turno": "MANHA",
            "status": "APROVADO",
            "responsavel": "Ana",
            "data": "14/06/2026",
            "observacao": "",
        }
    )
    retorno = {
        "lote_id": "LG-2026-00101",
        "screenshot": "logs/screenshots/playwright/LG-2026-00101.png",
        "driver": "playwright",
    }
    processar_playwright = Mock(return_value=retorno)

    with patch(
        "src.web_automation_playwright.processar_item_playwright",
        processar_playwright,
    ):
        resultado = web_automation.processar_item_web(item)

    assert resultado == retorno
    dados_lote = processar_playwright.call_args.args[0]
    assert dados_lote["lote_id"] == "LG-2026-00101"
    assert dados_lote["produto"] == "TV"
    assert dados_lote["status"] == "APROVADO"
    assert dados_lote["lote"] == "LG-2026-00101"


def test_montar_caminho_screenshot_cria_nome_seguro(monkeypatch, tmp_path):
    monkeypatch.setattr(web_evidencias, "SCREENSHOTS_DIR", tmp_path)

    caminho = montar_caminho_screenshot(
        {"lote_id": "LG/2026 00101"},
        "selenium",
    )

    assert "selenium_LG_2026_00101.png" in caminho.name
    assert caminho.parent.name == "selenium"
    assert caminho.parent.exists()
