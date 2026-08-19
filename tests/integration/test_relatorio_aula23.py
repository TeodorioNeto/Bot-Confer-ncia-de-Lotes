from datetime import datetime
from unittest.mock import MagicMock

import openpyxl
import pytest

import gerar_relatorio
from gerar_relatorio import (
    RegistroValidado,
    consolidar_e_validar,
    gerar_relatorio_excel,
)


@pytest.fixture
def registros_validados():
    return [
        RegistroValidado(
            aba_origem="Insp_01_08_2026",
            linha_planilha=4,
            data_referencia="2026-08-01",
            lote_id="LG-2026-00001",
            produto="AC12-SPLIT",
            linha="L1",
            turno="MANHA",
            status_original="APROVADO",
            status_normalizado="APROVADO",
            responsavel="Analista Teste",
            data_inspecao="01/08/2026",
            observacao="Sem divergencia",
            classificacao="Válido",
            regra_violada="-",
            motivo="Registro consistente com todas as regras RN01-RN12",
        ),
        RegistroValidado(
            aba_origem="Insp_01_08_2026",
            linha_planilha=5,
            data_referencia="2026-08-01",
            lote_id="LG-2026-99999",
            produto="AC12-SPLIT",
            linha="L1",
            turno="MANHA",
            status_original="APROVADO",
            status_normalizado="APROVADO",
            responsavel="Analista Teste",
            data_inspecao="01/08/2026",
            observacao="",
            classificacao="Divergência",
            regra_violada="RN05",
            motivo="Lote nao encontrado na Base_Referencia",
        ),
        RegistroValidado(
            aba_origem="Insp_01_08_2026",
            linha_planilha=6,
            data_referencia="2026-08-01",
            lote_id="LG-2026-00002",
            produto="MON27-QHD",
            linha="L2",
            turno="TARDE",
            status_original="APROVADO PARCIAL",
            status_normalizado="APROVADO PARCIAL",
            responsavel="Analista Teste",
            data_inspecao="01/08/2026",
            observacao="Revisar status",
            classificacao="Ambíguo",
            regra_violada="RN09",
            motivo="Status nao reconhecido",
        ),
        RegistroValidado(
            aba_origem="Insp_01_08_2026",
            linha_planilha=7,
            data_referencia="2026-08-01",
            lote_id=None,
            produto="TV55-4K-B",
            linha="L3",
            turno="NOITE",
            status_original="APROVADO",
            status_normalizado="APROVADO",
            responsavel="Analista Teste",
            data_inspecao="01/08/2026",
            observacao="",
            classificacao="Erro de Entrada",
            regra_violada="RN01-RN04",
            motivo="Campo obrigatorio vazio",
        ),
    ]


class DataFixa(datetime):
    @classmethod
    def now(cls):
        return cls(2026, 8, 17, 10, 30, 0)


@pytest.mark.integration
def test_gerar_relatorio_excel_cria_abas_resumo_graficos_e_dashboard(
    tmp_path,
    monkeypatch,
    registros_validados,
):
    monkeypatch.setattr(gerar_relatorio, "datetime", DataFixa)
    caminho_saida = tmp_path / "relatorio_conferencia_lotes.xlsx"

    gerar_relatorio_excel(registros_validados, caminho_saida)

    wb = openpyxl.load_workbook(caminho_saida)
    try:
        assert wb.sheetnames == [
            "Resumo",
            "Todos",
            "Válidos",
            "Divergências",
            "Ambíguos",
            "Erros de Entrada",
            "Ranking de Regras",
            "Dicionário",
            "Decisões de ML",
        ]
        assert len(wb["Resumo"]._charts) == 2
        assert wb["Resumo"]["B2"].value == 4
    finally:
        wb.close()


@pytest.mark.integration
@pytest.mark.regression
def test_consolidar_e_validar_usa_base_referencia_mockada(tmp_path, monkeypatch):
    caminho_entrada = tmp_path / "inspecao_lotes_dia.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insp_01_08_2026"
    ws.append(["Auditoria de lotes"])
    ws.append(["Gerado para teste"])
    ws.append(gerar_relatorio.COLUNAS_ENTRADA)
    ws.append(
        [
            "LG-2026-00001",
            "AC12-SPLIT",
            "L1",
            "MANHA",
            "APROVADO",
            "Analista Teste",
            "01/08/2026",
            "Sem divergencia",
        ]
    )
    ws.append(
        [
            "LG-2026-99999",
            "AC12-SPLIT",
            "L1",
            "MANHA",
            "APROVADO",
            "Analista Teste",
            "01/08/2026",
            "",
        ]
    )
    wb.create_sheet("Base_Referencia")
    wb.save(caminho_entrada)
    wb.close()

    carregar_base_mock = MagicMock(return_value={"LG-2026-00001"})
    monkeypatch.setattr(gerar_relatorio, "carregar_base_referencia", carregar_base_mock)

    registros = consolidar_e_validar(caminho_entrada)

    carregar_base_mock.assert_called_once_with(caminho_entrada)
    assert [registro.regra_violada for registro in registros] == ["-", "RN05"]
