from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

import gerar_relatorio as modulo_relatorio
from gerar_relatorio import (
    ARQUIVO_ENTRADA,
    RegistroValidado,
    classificar_registros_ambiguos_com_ml,
    gerar_relatorio_excel,
    gerar_resumo_executivo,
    gravar_decisoes_ml_em_excel,
    resolver_caminho_entrada,
)
from operational_indicators import consolidar_indicadores


ABAS_ESSENCIAIS = {
    "Resumo",
    "Todos",
    "Válidos",
    "Divergências",
    "Ambíguos",
    "Erros de Entrada",
    "Ranking de Regras",
    "Dicionário",
    "Decisões de ML",
}


def registro(classificacao, regra="-", linha_planilha=4):
    return RegistroValidado(
        aba_origem="Insp_01_08_2026",
        linha_planilha=linha_planilha,
        data_referencia="2026-08-01",
        lote_id=f"LG-2026-{linha_planilha:05d}",
        produto="AC12-SPLIT",
        linha="L1",
        turno="MANHA",
        status_original="APROVADO",
        status_normalizado="APROVADO",
        responsavel="Analista Teste",
        data_inspecao="01/08/2026",
        observacao="Sem divergencia",
        classificacao=classificacao,
        regra_violada=regra,
        motivo="Cenario controlado",
    )


@pytest.mark.integration
def test_relatorio_consolidado_cria_8_abas_essenciais_e_resumo_sincronizado(tmp_path):
    registros = [
        registro("Válido", linha_planilha=4),
        registro("Divergência", "RN05", linha_planilha=5),
        registro("Divergência", "RN05", linha_planilha=6),
        registro("Ambíguo", "RN09", linha_planilha=7),
        registro("Erro de Entrada", "RN12", linha_planilha=8),
    ]
    indicadores = consolidar_indicadores(registros)
    caminho_excel = tmp_path / "relatorio_conferencia_lotes.xlsx"
    caminho_md = tmp_path / "resumo_executivo.md"

    decisoes_ml = [
        {
            "lote_id": "LG-2026-00007",
            "classe": "revisar",
            "probabilidade": 0.72,
            "nivel_confianca": "media",
            "acao": "revisar",
            "latencia_ms": 18.4,
            "fallback": False,
        }
    ]

    gerar_relatorio_excel(registros, caminho_excel, indicadores, decisoes_ml)
    gerar_resumo_executivo(indicadores, caminho_md)

    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    try:
        assert set(wb.sheetnames) == ABAS_ESSENCIAIS
        assert wb["Resumo"]["B2"].value == indicadores.total_registros
        assert wb["Resumo"]["B7"].value.startswith("RN05")
        assert wb["Ranking de Regras"]["A2"].value == "RN05"
        assert wb["Ranking de Regras"]["C2"].value == 2
        assert wb["Dicionário"]["A2"].value == "Válido"
        assert wb["Decisões de ML"]["A2"].value == "LG-2026-00007"
        assert wb["Decisões de ML"]["B2"].value == "revisar"
        assert len(wb["Resumo"]._charts) == 2
    finally:
        wb.close()

    resumo = caminho_md.read_text(encoding="utf-8")
    assert f"Foram processados {indicadores.total_registros} registros" in resumo
    assert "RN05" in resumo
    assert f"{indicadores.ganho_estimado_minutos:.0f} minutos" in resumo


@pytest.mark.integration
def test_resolver_entrada_padrao_aceita_arquivo_legado_na_raiz(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    caminho_legado = tmp_path / "inspecao_lotes_10dias.xlsx"
    caminho_legado.touch()

    assert resolver_caminho_entrada(ARQUIVO_ENTRADA) == Path(caminho_legado.name)


@pytest.mark.integration
def test_atualiza_aba_decisoes_ml_sem_perder_relatorio_existente(tmp_path):
    registros = [registro("Ambíguo", "RN09", linha_planilha=7)]
    caminho_excel = tmp_path / "relatorio.xlsx"
    gerar_relatorio_excel(registros, caminho_excel)

    decisoes_ml = [
        {
            "lote_id": "LG-2026-00007",
            "classe": "REVISAO_ML_OFFLINE",
            "probabilidade": 0.0,
            "nivel_confianca": "offline",
            "acao": "revisar",
            "latencia_ms": 0.0,
            "fallback": True,
        }
    ]
    gravar_decisoes_ml_em_excel(caminho_excel, decisoes_ml)

    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    try:
        assert wb["Resumo"]["B2"].value == 1
        assert len(wb["Resumo"]._charts) == 2
        assert wb["Decisões de ML"]["A2"].value == "LG-2026-00007"
        assert wb["Decisões de ML"]["B2"].value == "REVISAO_ML_OFFLINE"
        assert wb["Decisões de ML"].max_row == 2
    finally:
        wb.close()


@pytest.mark.integration
def test_classifica_somente_ambiguos_da_planilha_consolidada():
    valido = registro("Válido", linha_planilha=4)
    ambiguo_a = registro("Ambíguo", "RN09", linha_planilha=7)
    ambiguo_a.status_original = "APROVADO PARCIAL"
    ambiguo_a.turno = "B"
    ambiguo_b = registro("Ambíguo", "RN09", linha_planilha=8)
    ambiguo_b.status_original = "EM AJUSTE"
    ambiguo_b.turno = "C"

    class ClientFake:
        def __init__(self):
            self.payloads = []

        def classificar(self, payload):
            self.payloads.append(payload)
            return SimpleNamespace(
                lote_id=payload["lote_id"],
                classe="revisar",
                probabilidade=0.91,
                nivel_confianca="alta",
                acao="acao_automatica",
                latencia_ms=4.2,
            )

    client = ClientFake()
    decisoes = classificar_registros_ambiguos_com_ml(
        [valido, ambiguo_a, ambiguo_b],
        client,
    )

    assert len(decisoes) == 2
    assert len(client.payloads) == 2
    assert client.payloads[0]["turno"] == "TARDE"
    assert client.payloads[1]["turno"] == "NOITE"
    assert all(decisao["fallback"] is False for decisao in decisoes)


@pytest.mark.integration
def test_main_do_relatorio_gera_decisoes_dos_ambiguos_da_mesma_planilha(
    tmp_path,
    monkeypatch,
):
    capturado = {}
    registros = [registro("Ambíguo", "RN09", linha_planilha=7)]
    monkeypatch.setattr(
        modulo_relatorio,
        "consolidar_e_validar",
        lambda caminho: registros,
    )

    def gerar_excel(registros_recebidos, caminho, indicadores, decisoes_recebidas):
        capturado["decisoes_ml"] = decisoes_recebidas
        return Path(caminho)

    monkeypatch.setattr(modulo_relatorio, "gerar_relatorio_excel", gerar_excel)
    monkeypatch.setattr(
        modulo_relatorio,
        "gerar_resumo_executivo",
        lambda indicadores, caminho: Path(caminho),
    )

    class ClientFake:
        def classificar(self, payload):
            return SimpleNamespace(
                lote_id=payload["lote_id"],
                classe="revisar",
                probabilidade=0.72,
                nivel_confianca="media",
                acao="revisar",
                latencia_ms=18.4,
            )

    modulo_relatorio.main(
        caminho_entrada=tmp_path / "entrada.xlsx",
        caminho_saida=tmp_path / "relatorio.xlsx",
        caminho_resumo=tmp_path / "resumo.md",
        ml_client=ClientFake(),
    )

    assert len(capturado["decisoes_ml"]) == 1
    assert capturado["decisoes_ml"][0]["lote_id"] == "LG-2026-00007"
    assert capturado["decisoes_ml"][0]["fallback"] is False
