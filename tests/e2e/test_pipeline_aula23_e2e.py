from datetime import datetime
from unittest.mock import MagicMock

import openpyxl
import pytest

import gerar_relatorio
from gerar_relatorio import consolidar_e_validar, gerar_relatorio_excel


class DataFixa(datetime):
    @classmethod
    def now(cls):
        return cls(2026, 8, 17, 11, 0, 0)


@pytest.mark.e2e
def test_pipeline_completo_com_planilha_temporaria(tmp_path, monkeypatch):
    caminho_entrada = tmp_path / "inspecao_lotes_dia.xlsx"
    caminho_saida = tmp_path / "relatorio_conferencia_lotes.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insp_01_08_2026"
    ws.append(["Auditoria de lotes"])
    ws.append(["Gerado para teste E2E"])
    ws.append(gerar_relatorio.COLUNAS_ENTRADA)
    ws.append(
        [
            "LG-2026-00001",
            "AC12-SPLIT",
            "L1",
            "MANHA",
            "APROVADO",
            "Analista Teste",
            "01/08/2026",
            "Sem divergencia",
        ]
    )
    ws.append(
        [
            "LG-2026-00002",
            "MON27-QHD",
            "L2",
            "TARDE",
            "NOK",
            "Analista Teste",
            "01/08/2026",
            "",
        ]
    )
    wb.create_sheet("Base_Referencia")
    wb.save(caminho_entrada)
    wb.close()

    monkeypatch.setattr(
        gerar_relatorio,
        "carregar_base_referencia",
        MagicMock(return_value={"LG-2026-00001", "LG-2026-00002"}),
    )
    monkeypatch.setattr(gerar_relatorio, "datetime", DataFixa)

    registros = consolidar_e_validar(caminho_entrada)
    gerar_relatorio_excel(registros, caminho_saida)

    assert [registro.regra_violada for registro in registros] == ["-", "RN10"]
    assert caminho_saida.exists()

    relatorio = openpyxl.load_workbook(caminho_saida, data_only=True)
    try:
        assert relatorio.sheetnames == [
            "Resumo",
            "Todos",
            "Válidos",
            "Divergências",
            "Ambíguos",
            "Erros de Entrada",
            "Ranking de Regras",
            "Dicionário",
            "Decisões de ML",
        ]
        assert relatorio["Resumo"]["B2"].value == 2
        assert relatorio["Ranking de Regras"]["A2"].value == "RN10"
    finally:
        relatorio.close()
