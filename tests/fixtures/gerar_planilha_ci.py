from pathlib import Path

import openpyxl


COLUNAS = [
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
]


def main():
    destino = Path("dados_entrada") / "inspecao_lotes_dia.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["INSPECAO DE LOTES - MASSA SINTETICA CI"])
    ws.append(["Gerada automaticamente para Docker/CI"])
    ws.append(COLUNAS)
    ws.append(
        [
            "LG-2026-91001",
            "TV55-4K-B",
            "L1",
            "MANHA",
            "APROVADO",
            "Usuario CI",
            "2026-07-14",
            "",
        ]
    )
    ws.append(
        [
            "LG-2026-91002",
            "AC12-SPLIT",
            "L2",
            "TARDE",
            "REPROVADO",
            "Usuario CI",
            "2026-07-14",
            "",
        ]
    )

    base = wb.create_sheet("Base_Referencia")
    base.append(["BASE DE REFERENCIA DE LOTES - MASSA SINTETICA CI"])
    base.append(["lote_id"])
    base.append(["LG-2026-91001"])
    base.append(["LG-2026-91002"])

    formulario = wb.create_sheet("Formulario_Analise")
    formulario.append(["FORMULARIO DE ANALISE - MASSA SINTETICA CI"])
    formulario.append(["Gerado automaticamente para Docker/CI"])
    formulario.append(
        [
            "registro",
            "lote_id",
            "problema",
            "regra",
            "acao_recomendada",
            "revisao",
        ]
    )

    wb.create_sheet("lotes_ambiguos")
    wb.create_sheet("Resumo_Diario")

    wb.save(destino)


if __name__ == "__main__":
    main()
