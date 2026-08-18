import openpyxl

import dispatcher
from bot import processar_item


class FakeDataPool:
    def __init__(self, tem_pendente=False):
        self.tem_pendente = tem_pendente
        self.entries = []
        self.cursor = 0

    def has_next(self):
        return self.tem_pendente or self.cursor < len(self.entries)

    def create_entry(self, entry):
        self.entries.append(entry)

    def next(self, task_id=None):
        if self.cursor >= len(self.entries):
            return None

        item = self.entries[self.cursor]
        self.cursor += 1
        return item


class FakeMaestro:
    def __init__(self, datapool):
        self.datapool = datapool

    def get_datapool(self, label):
        self.label = label
        return self.datapool


class FakeDataPoolEntry:
    def __init__(self, values):
        self.values = values
        self.status_reportado = None
        self.mensagem = ""

    def get_value(self, chave):
        return self.values.get(chave)

    def report_done(self, finish_message=""):
        self.status_reportado = "DONE"
        self.mensagem = finish_message

    def report_error(self, error_type=None, finish_message=""):
        self.status_reportado = "ERROR"
        self.mensagem = finish_message


def criar_planilha_inspecao(caminho):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["PLANILHA DE INSPECAO"])
    ws.append(["Arquivo", "Sistema", "Registros"])
    ws.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
        ]
    )
    ws.append(["LG-2026-00101", "TV", "A", "MANHA", "APROVADO", "Ana", "14/06/2026", ""])
    ws.append(["Legenda", None, None, None, None, None, None, None])
    wb.save(caminho)


def test_popular_fila_publica_lotes_validos_e_ignora_rodape(monkeypatch, tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    criar_planilha_inspecao(caminho)
    datapool = FakeDataPool()

    monkeypatch.setattr(dispatcher, "ARQUIVO_INSPECAO", caminho)
    monkeypatch.setattr(dispatcher, "DataPoolEntry", FakeDataPoolEntry)
    monkeypatch.setattr(dispatcher, "valida_estrutura", lambda _: True)

    resultado = dispatcher.popular_fila(FakeMaestro(datapool))

    assert resultado == {
        "enviados": 1,
        "ignorados": 1,
        "fila_ja_populada": False,
    }
    assert len(datapool.entries) == 1
    assert datapool.entries[0].values["lote_id"] == "LG-2026-00101"
    assert datapool.entries[0].values["linha_planilha"] == 4
    assert datapool.entries[0].values["screenshot"] == ""


def test_popular_fila_nao_republica_quando_ja_tem_item_pendente(monkeypatch, tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    criar_planilha_inspecao(caminho)
    datapool = FakeDataPool(tem_pendente=True)

    monkeypatch.setattr(dispatcher, "ARQUIVO_INSPECAO", caminho)

    resultado = dispatcher.popular_fila(FakeMaestro(datapool))

    assert resultado == {
        "enviados": 0,
        "ignorados": 0,
        "fila_ja_populada": True,
    }
    assert datapool.entries == []


def test_dispatcher_e_performer_processam_todos_os_itens_do_datapool(monkeypatch, tmp_path):
    caminho = tmp_path / "inspecao_lotes_dia.xlsx"
    criar_planilha_inspecao(caminho)
    wb = openpyxl.load_workbook(caminho)
    ws = wb["Inspecao_14_06_2026"]
    ws.insert_rows(5)
    ws.cell(row=5, column=1, value="LG-2026-00103")
    ws.cell(row=5, column=2, value="TV")
    ws.cell(row=5, column=3, value="A")
    ws.cell(row=5, column=4, value="MANHA")
    ws.cell(row=5, column=5, value="APROVADO")
    ws.cell(row=5, column=6, value="Bia")
    ws.cell(row=5, column=7, value="14/06/2026")
    ws.cell(row=5, column=8, value="")
    wb.save(caminho)
    wb.close()

    datapool = FakeDataPool()
    monkeypatch.setattr(dispatcher, "ARQUIVO_INSPECAO", caminho)
    monkeypatch.setattr(dispatcher, "DataPoolEntry", FakeDataPoolEntry)
    monkeypatch.setattr(dispatcher, "valida_estrutura", lambda _: True)

    dispatcher.popular_fila(FakeMaestro(datapool))

    resultados = []
    while datapool.has_next():
        item = datapool.next()
        resultado = processar_item(item, {"LG-2026-00101"})
        resultados.append(resultado)

        if resultado["divergencias"]:
            item.report_error(finish_message=" | ".join(resultado["divergencias"]))
        else:
            item.report_done()

    assert [item.status_reportado for item in datapool.entries] == ["DONE", "ERROR"]
    assert resultados[0]["lote_id"] == "LG-2026-00101"
    assert resultados[1]["lote_id"] == "LG-2026-00103"
    assert resultados[1]["divergencias"]
