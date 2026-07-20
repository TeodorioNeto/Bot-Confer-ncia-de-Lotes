import openpyxl

from src.analise_formulario import analisar_e_preencher_formulario


def criar_workbook_inspecao(caminho):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Inspecao_14_06_2026"
    ws.append(["PLANILHA DE INSPECAO DE LOTES"])
    ws.append(["Arquivo", "Sistema", "Registros"])
    ws.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
        ]
    )
    ws.append(["LG-2026-00101", "TV", "A", "MANHA", "APROVADO", "Ana", "14/06/2026", ""])
    ws.append(
        [
            "LG-2026-00102",
            "TV",
            "A",
            "MANHA",
            "APROVADO PARCIAL",
            "Bruno",
            "14/06/2026",
            "Aguardando revisao",
        ]
    )
    ws.append(["LG-2026-00103", "TV", "A", "MANHA", "NOK", "Carla", "14/06/2026", ""])

    analise = wb.create_sheet("Formulario_Analise")
    for _ in range(40):
        analise.append(["", "", "", "", "", ""])

    base = wb.create_sheet("Base_Referencia")
    base.append(["BASE DE REFERENCIA"])
    base.append(["lote_id", "codigo_produto", "descricao_produto", "status_cadastro"])
    base.append(["LG-2026-00101", "P001", "TV", "ativo"])
    base.append(["LG-2026-00102", "P002", "TV", "ativo"])
    base.append(["LG-2026-00103", "P003", "TV", "ativo"])

    wb.save(caminho)


def test_planilha_final_cria_lotes_ambiguos_e_resumo_diario(tmp_path):
    entrada = tmp_path / "entrada.xlsx"
    saida = tmp_path / "saida.xlsx"
    criar_workbook_inspecao(entrada)

    caminho, resultados, resumo = analisar_e_preencher_formulario(entrada, saida)

    assert caminho == saida
    assert len(resultados) == 3
    assert resumo["lotes_ambiguos"] == 1

    wb = openpyxl.load_workbook(saida, data_only=False)
    assert "Formulario_Analise" in wb.sheetnames
    assert "lotes_ambiguos" in wb.sheetnames
    assert "Resumo_Diario" in wb.sheetnames

    ambiguos = wb["lotes_ambiguos"]
    assert ambiguos["A1"].value == "registro"
    assert ambiguos["E2"].value == "RN06"
    assert ambiguos["D2"].value == "APROVADO PARCIAL"

    resumo_diario = wb["Resumo_Diario"]
    indicadores = {
        resumo_diario.cell(row=linha, column=1).value: resumo_diario.cell(row=linha, column=2).value
        for linha in range(2, resumo_diario.max_row + 1)
    }
    assert indicadores["total_registros"] == 3
    assert indicadores["lotes_ambiguos"] == 1
    assert indicadores["registros_com_divergencia"] == 2
