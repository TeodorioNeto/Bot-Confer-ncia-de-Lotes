"""
gerar_relatorio.py

Escopo deste script:
- Etapas 5.1 e 5.2: consolidar, validar (RN01-RN12) e gerar as 8 abas do Excel.
- Etapa 5.3: montar o dashboard executivo na aba 'Resumo' com gráficos nativos.
- Etapa 5.4: Gerar saidas executivas sem misturar logs ao Excel.
- Aula 24-A: adicionar a aba 'Decisões de ML' para auditoria do classificador.
"""

from __future__ import annotations

import re
from collections import Counter
from datetime import datetime
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, LineChart, Reference
from operational_indicators import (
    OperationalIndicators,
    REGRA_NOMES,
    consolidar_indicadores,
)
from config import ML_API_URL, ML_MAX_FAILURES, ML_TIMEOUT_SECONDS
from item_processor import classificar_ambiguo_com_ml
from src.ml_client import MLClient

ARQUIVO_ENTRADA = "dados_entrada/inspecao_lotes_10dias.xlsx"
ARQUIVO_ENTRADA_LEGADO = "inspecao_lotes_10dias.xlsx"
ARQUIVO_SAIDA = "relatorio_conferencia_lotes.xlsx"
ARQUIVO_RESUMO_EXECUTIVO = "resumo_executivo.md"

COLUNAS_DECISOES_ML = [
    "lote_id",
    "classe",
    "probabilidade",
    "nivel_confianca",
    "acao",
    "latencia_ms",
    "fallback",
]

COLUNAS_ENTRADA = [
    "lote_id", "produto", "linha", "turno",
    "status", "responsavel", "data", "observacao",
]

STATUS_PADRAO = {"APROVADO", "REPROVADO", "PENDENTE"}
SINONIMOS_STATUS = {"OK": "APROVADO", "NOK": "REPROVADO"}
STATUS_REPROVADO_ORIGINAL = {"REPROVADO", "NOK"}

# RN12: data no formato DD/MM/AAAA
DATA_REGEX = re.compile(r"^\d{2}/\d{2}/\d{4}$")


# ---------------------------------------------------------------------------
# RegistroValidado
# ---------------------------------------------------------------------------

@dataclass
class RegistroValidado:
    aba_origem: str
    linha_planilha: int
    data_referencia: str            # data do dia da coleta (nome da aba), AAAA-MM-DD
    lote_id: Optional[str]
    produto: Optional[str]
    linha: Optional[str]
    turno: Optional[str]
    status_original: Optional[str]
    status_normalizado: Optional[str]
    responsavel: Optional[str]
    data_inspecao: Optional[str]    # valor originalmente digitado na coluna "data"
    observacao: Optional[str]
    classificacao: str              # Válido | Divergência | Ambíguo | Erro de Entrada
    regra_violada: str              # ex.: "RN05", "RN01-RN04", "-"
    motivo: str

    def to_dict(self) -> dict:
        return asdict(self)


# ---------------------------------------------------------------------------
# Regras de negócio RN01-RN12
# ---------------------------------------------------------------------------

def vazio(valor) -> bool:
    return valor is None or str(valor).strip() == ""


def normalizar_status(status) -> Optional[str]:
    """RN06-RN07: 'OK' -> APROVADO, 'NOK' -> REPROVADO. Demais valores, upper/strip."""
    if vazio(status):
        return None
    s = str(status).strip().upper()
    return SINONIMOS_STATUS.get(s, s)


def data_valida(data_texto) -> bool:
    """RN12: data ausente ou fora do formato DD/MM/AAAA é inválida."""
    if vazio(data_texto):
        return False
    texto = str(data_texto).strip()
    if not DATA_REGEX.match(texto):
        return False
    try:
        datetime.strptime(texto, "%d/%m/%Y")
        return True
    except ValueError:
        return False


def validar_registro(registro: dict, ocorrencia_no_dia: int, base_referencia: set) -> RegistroValidado:
    produto = registro.get("produto")
    linha = registro.get("linha")
    status_original = registro.get("status")
    data_inspecao = registro.get("data")
    observacao = registro.get("observacao")

    status_normalizado = normalizar_status(status_original)

    # RN01-RN04
    campos_vazios = [
        nome for nome, valor in [
            ("lote_id", registro.get("lote_id")), ("produto", produto),
            ("linha", linha), ("status", status_original),
        ] if vazio(valor)
    ]
    if campos_vazios:
        return _montar_registro(
            registro, status_original, status_normalizado,
            classificacao="Erro de Entrada",
            regra_violada="RN01-RN04",
            motivo=f"Campo(s) obrigatório(s) vazio(s): {', '.join(campos_vazios)}",
        )

    # RN12
    if not data_valida(data_inspecao):
        return _montar_registro(
            registro, status_original, status_normalizado,
            classificacao="Erro de Entrada",
            regra_violada="RN12",
            motivo=f"Data de inspeção ausente ou fora do formato DD/MM/AAAA: '{data_inspecao}'",
        )

    # RN11
    if ocorrencia_no_dia >= 2:
        return _montar_registro(
            registro, status_original, status_normalizado,
            classificacao="Divergência",
            regra_violada="RN11",
            motivo=f"Lote duplicado no mesmo dia (ocorrência nº {ocorrencia_no_dia})",
        )

    # RN09
    if status_normalizado not in STATUS_PADRAO:
        return _montar_registro(
            registro, status_original, status_normalizado,
            classificacao="Ambíguo",
            regra_violada="RN09",
            motivo=f"Status '{status_original}' não reconhecido — necessita revisão humana",
        )

    # RN05
    lote_id_str = str(registro.get("lote_id")).strip()
    if lote_id_str not in base_referencia:
        return _montar_registro(
            registro, status_original, status_normalizado,
            classificacao="Divergência",
            regra_violada="RN05",
            motivo=f"Lote '{lote_id_str}' não encontrado na Base_Referencia",
        )

    # RN10
    if str(status_original).strip().upper() in STATUS_REPROVADO_ORIGINAL and vazio(observacao):
        return _montar_registro(
            registro, status_original, status_normalizado,
            classificacao="Divergência",
            regra_violada="RN10",
            motivo="Lote reprovado sem observação preenchida",
        )

    # RN08
    return _montar_registro(
        registro, status_original, status_normalizado,
        classificacao="Válido",
        regra_violada="-",
        motivo="Registro consistente com todas as regras RN01-RN12",
    )


def _montar_registro(registro, status_original, status_normalizado, classificacao, regra_violada, motivo) -> RegistroValidado:
    lote_id = registro.get("lote_id")
    return RegistroValidado(
        aba_origem=registro["aba_origem"],
        linha_planilha=registro["linha_planilha"],
        data_referencia=registro["data_referencia"],
        lote_id=None if vazio(lote_id) else str(lote_id).strip(),
        produto=registro.get("produto"),
        linha=registro.get("linha"),
        turno=registro.get("turno"),
        status_original=status_original,
        status_normalizado=status_normalizado,
        responsavel=registro.get("responsavel"),
        data_inspecao=registro.get("data"),
        observacao=registro.get("observacao"),
        classificacao=classificacao,
        regra_violada=regra_violada,
        motivo=motivo,
    )


# ---------------------------------------------------------------------------
# Etapa 5.1 - Leitura, deduplicação e consolidação
# ---------------------------------------------------------------------------

def carregar_base_referencia(caminho) -> set:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Base_Referencia"]
    base = set()
    for linha in ws.iter_rows(min_row=3, values_only=True):
        lote_id = linha[0] if linha else None
        if lote_id and str(lote_id).strip().startswith("LG-"):
            base.add(str(lote_id).strip())
    wb.close()
    return base


def nome_aba_para_data(nome_aba: str) -> str:
    dia, mes, ano = nome_aba.replace("Insp_", "").split("_")
    return f"{ano}-{mes}-{dia}"


def ler_aba_diaria(ws, nome_aba: str) -> list[dict]:
    data_referencia = nome_aba_para_data(nome_aba)
    registros = []
    for linha_excel, linha in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        valores = dict(zip(COLUNAS_ENTRADA, (linha or ())[: len(COLUNAS_ENTRADA)]))
        if all(v is None or str(v).strip() == "" for v in valores.values()):
            continue
        lote_id_bruto = valores.get("lote_id")
        if not vazio(lote_id_bruto) and str(lote_id_bruto).strip().lower().startswith("total"):
            continue
        valores["aba_origem"] = nome_aba
        valores["linha_planilha"] = linha_excel
        valores["data_referencia"] = data_referencia
        registros.append(valores)
    return registros


def resolver_caminho_entrada(caminho_entrada) -> Path:
    """Resolve a entrada padrão e mantém compatibilidade com o layout legado."""
    caminho = Path(caminho_entrada)
    if caminho.is_file():
        return caminho

    caminho_padrao = Path(ARQUIVO_ENTRADA)
    caminho_legado = Path(ARQUIVO_ENTRADA_LEGADO)
    if caminho == caminho_padrao and caminho_legado.is_file():
        return caminho_legado

    tentativas = [caminho]
    if caminho == caminho_padrao:
        tentativas.append(caminho_legado)
    caminhos = ", ".join(str(item) for item in tentativas)
    raise FileNotFoundError(f"Arquivo de entrada não encontrado. Caminhos verificados: {caminhos}")


def consolidar_e_validar(caminho_entrada) -> list[RegistroValidado]:
    caminho_entrada = resolver_caminho_entrada(caminho_entrada)
    wb = openpyxl.load_workbook(caminho_entrada, data_only=True)
    base_referencia = carregar_base_referencia(caminho_entrada)
    abas_diarias = [nome for nome in wb.sheetnames if nome != "Base_Referencia"]

    todos_validados: list[RegistroValidado] = []
    for nome_aba in abas_diarias:
        registros_do_dia = ler_aba_diaria(wb[nome_aba], nome_aba)

        contador = Counter()
        for registro in registros_do_dia:
            lote_id = registro.get("lote_id")
            if vazio(lote_id):
                ocorrencia = 1
            else:
                chave = str(lote_id).strip()
                contador[chave] += 1
                ocorrencia = contador[chave]

            todos_validados.append(validar_registro(registro, ocorrencia, base_referencia))

    wb.close()
    return todos_validados


# ---------------------------------------------------------------------------
# Etapas 5.2, 5.3 e 5.4 - Relatório, Dashboard e Log
# ---------------------------------------------------------------------------

RENOMEAR_COLUNAS = {
    "data_referencia": "Data",
    "lote_id": "Lote",
    "produto": "Produto",
    "linha": "Linha de Produção",
    "turno": "Turno",
    "status_normalizado": "Status",
    "responsavel": "Responsável",
    "observacao": "Observação",
    "classificacao": "Classificação",
    "regra_violada": "Regra",
    "motivo": "Motivo",
    "aba_origem": "Aba de Origem",
    "status_original": "Status Informado",
    "data_inspecao": "Data Informada",
    "linha_planilha": "Linha na Planilha",
}

COLUNAS_RELATORIO = [
    "data_referencia", "lote_id", "produto", "linha", "turno",
    "status_normalizado", "status_original", "responsavel",
    "observacao", "classificacao", "regra_violada", "motivo",
    "aba_origem", "linha_planilha",
]


def montar_dataframe(registros: list[RegistroValidado]) -> pd.DataFrame:
    df = pd.DataFrame([r.to_dict() for r in registros])
    df = df[COLUNAS_RELATORIO].rename(columns=RENOMEAR_COLUNAS)
    df = df.sort_values(["Data", "Lote"], na_position="last").reset_index(drop=True)
    return df


def montar_resumo(indicadores: OperationalIndicators) -> pd.DataFrame:
    linhas = [
        {
            "Indicador": "Total de registros",
            "Valor": indicadores.total_registros,
            "%": "-",
            "Referencia": "-",
            "Sinal": "-",
        },
        {
            "Indicador": "Registros válidos",
            "Valor": indicadores.registros_validos,
            "%": _fmt_pct(indicadores.percentual_validos),
            "Referencia": "Informativa",
            "Sinal": "-",
        },
        {
            "Indicador": "Divergências",
            "Valor": indicadores.divergencias,
            "%": _fmt_pct(indicadores.percentual_divergencias),
            "Referencia": "Informativa",
            "Sinal": "-",
        },
        {
            "Indicador": "Ambíguos",
            "Valor": indicadores.ambiguos,
            "%": _fmt_pct(indicadores.percentual_ambiguos),
            "Referencia": "Informativa",
            "Sinal": "-",
        },
        {
            "Indicador": "Erros de Entrada",
            "Valor": indicadores.erros_entrada,
            "%": _fmt_pct(indicadores.percentual_erros_entrada),
            "Referencia": "Informativa",
            "Sinal": "-",
        },
        {
            "Indicador": "Regra mais acionada",
            "Valor": (
                f"{indicadores.regra_mais_acionada} - "
                f"{indicadores.regra_mais_acionada_nome} "
                f"({indicadores.regra_mais_acionada_quantidade})"
            ),
            "%": "-",
            "Referencia": "-",
            "Sinal": "-",
        },
        {
            "Indicador": "Taxa de qualidade da entrada",
            "Valor": _fmt_pct(indicadores.taxa_qualidade_entrada),
            "%": _fmt_pct(indicadores.taxa_qualidade_entrada),
            "Referencia": "> 80%",
            "Sinal": _sinal_meta(indicadores.taxa_qualidade_entrada >= 80),
        },
        {
            "Indicador": "Taxa de revisão humana",
            "Valor": _fmt_pct(indicadores.taxa_revisao_humana),
            "%": _fmt_pct(indicadores.taxa_revisao_humana),
            "Referencia": "< 15%",
            "Sinal": _sinal_meta(indicadores.taxa_revisao_humana < 15),
        },
        {
            "Indicador": "Taxa de retrabalho",
            "Valor": _fmt_pct(indicadores.taxa_retrabalho),
            "%": _fmt_pct(indicadores.taxa_retrabalho),
            "Referencia": "< 6%",
            "Sinal": _sinal_meta(indicadores.taxa_retrabalho < 6),
        },
        {
            "Indicador": "Ganho estimado de tempo",
            "Valor": (
                f"{indicadores.ganho_estimado_minutos:.0f} min "
                f"({indicadores.ganho_estimado_horas:.1f} h)"
            ),
            "%": "-",
            "Referencia": "Estimativa didática",
            "Sinal": "-",
        },
    ]
    return pd.DataFrame(linhas)


def montar_ranking_regras(indicadores: OperationalIndicators) -> pd.DataFrame:
    linhas = [
        {
            "Regra": item.regra,
            "Descrição": item.nome,
            "Ocorrências": item.quantidade,
            "% do total": _fmt_pct(item.percentual_total),
        }
        for item in indicadores.ranking_regras
    ]
    if not linhas:
        linhas.append(
            {
                "Regra": "-",
                "Descrição": "Nenhuma regra acionada",
                "Ocorrências": 0,
                "% do total": "0,0%",
            }
        )
    return pd.DataFrame(linhas)


def montar_dicionario() -> pd.DataFrame:
    termos = [
        ("Válido", "Registro que passou por todas as validações automáticas."),
        ("Divergência", "Registro que precisa de reconciliação com a base ou o processo."),
        ("Ambíguo", "Registro que precisa de revisão humana por status não reconhecido."),
        ("Erro de Entrada", "Registro com falha básica de preenchimento ou data inválida."),
        ("RN05", REGRA_NOMES["RN05"]),
        ("RN09", REGRA_NOMES["RN09"]),
        ("RN10", REGRA_NOMES["RN10"]),
        ("RN11", REGRA_NOMES["RN11"]),
        ("RN12", REGRA_NOMES["RN12"]),
        ("Taxa de qualidade da entrada", "Percentual de registros sem erro básico de entrada."),
        ("Taxa de revisão humana", "Percentual de registros que a automação não decide sozinha."),
        ("Taxa de retrabalho", "Percentual de registros que precisam de reconciliação."),
        ("Ganho estimado de tempo", "Estimativa didática baseada em premissas de tempo manual e automatizado."),
    ]
    return pd.DataFrame(termos, columns=["Termo", "Descrição"])


def montar_decisoes_ml(decisoes_ml: list[dict] | None = None) -> pd.DataFrame:
    return pd.DataFrame(decisoes_ml or [], columns=COLUNAS_DECISOES_ML)


def formatar_aba(ws, cor_cabecalho="1F4E78"):
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for coluna in ws.columns:
        maior = max((len(str(c.value)) for c in coluna if c.value is not None), default=10)
        letra = get_column_letter(coluna[0].column)
        ws.column_dimensions[letra].width = min(maior + 2, 50)


def gravar_decisoes_ml_em_excel(
    caminho_excel,
    decisoes_ml: list[dict] | None,
) -> Path:
    """Cria ou substitui a aba de auditoria ML em um relatório existente."""
    caminho_excel = Path(caminho_excel)
    if not caminho_excel.exists():
        raise FileNotFoundError(f"Relatório Excel não encontrado: {caminho_excel}")

    wb = openpyxl.load_workbook(caminho_excel)
    try:
        if "Decisões de ML" in wb.sheetnames:
            wb.remove(wb["Decisões de ML"])

        ws = wb.create_sheet("Decisões de ML")
        ws.append(COLUNAS_DECISOES_ML)
        for decisao in decisoes_ml or []:
            ws.append([decisao.get(coluna) for coluna in COLUNAS_DECISOES_ML])

        formatar_aba(ws)
        wb.save(caminho_excel)
    finally:
        wb.close()

    return caminho_excel


def classificar_registros_ambiguos_com_ml(
    registros: list[RegistroValidado],
    ml_client,
) -> list[dict]:
    """Classifica somente os ambíguos consolidados da planilha de 10 dias."""
    decisoes_ml = []
    for registro in registros:
        if registro.classificacao != "Ambíguo":
            continue

        item_ml = {
            "lote_id": registro.lote_id,
            "status": registro.status_original,
            "turno": registro.turno,
            "observacao": registro.observacao,
        }
        decisoes_ml.append(classificar_ambiguo_com_ml(item_ml, ml_client))

    return decisoes_ml


def _fmt_pct(valor: float) -> str:
    return f"{valor:.1f}%".replace(".", ",")


def _fmt_decimal(valor: float) -> str:
    return f"{valor:.1f}".replace(".", ",")


def _sinal_meta(ok: bool) -> str:
    return "OK" if ok else "ATENCAO"


def _plural(quantidade: int, singular: str, plural: str) -> str:
    return singular if quantidade == 1 else plural


def gerar_relatorio_excel(
    registros: list[RegistroValidado],
    caminho_saida,
    indicadores: OperationalIndicators | None = None,
    decisoes_ml: list[dict] | None = None,
) -> Path:
    indicadores = indicadores or consolidar_indicadores(registros)
    df_todos = montar_dataframe(registros)
    df_validos = df_todos[df_todos["Classificação"] == "Válido"]
    df_divergencias = df_todos[df_todos["Classificação"] == "Divergência"]
    df_ambiguos = df_todos[df_todos["Classificação"] == "Ambíguo"]
    df_erros = df_todos[df_todos["Classificação"] == "Erro de Entrada"]
    df_resumo = montar_resumo(indicadores)
    df_ranking = montar_ranking_regras(indicadores)
    df_dicionario = montar_dicionario()
    df_decisoes_ml = montar_decisoes_ml(decisoes_ml)

    caminho_saida = Path(caminho_saida)
    
    # 1. Escrita inicial das abas com pandas[cite: 1]
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_todos.to_excel(writer, sheet_name="Todos", index=False)
        df_validos.to_excel(writer, sheet_name="Válidos", index=False)
        df_divergencias.to_excel(writer, sheet_name="Divergências", index=False)
        df_ambiguos.to_excel(writer, sheet_name="Ambíguos", index=False)
        df_erros.to_excel(writer, sheet_name="Erros de Entrada", index=False)
        df_ranking.to_excel(writer, sheet_name="Ranking de Regras", index=False)
        df_dicionario.to_excel(writer, sheet_name="Dicionário", index=False)
        df_decisoes_ml.to_excel(writer, sheet_name="Decisões de ML", index=False)

    # 2. Pós-processamento com openpyxl (Etapas 3 e 4)[cite: 1]
    wb = openpyxl.load_workbook(caminho_saida)
    ws_resumo = wb["Resumo"]

    for nome_aba in wb.sheetnames:
        formatar_aba(wb[nome_aba])

    # --- ETAPA 5.3: Gráficos Nativos no Dashboard (Aba Resumo)[cite: 1] ---
    # A) Gráfico de Rosca
    donut = DoughnutChart()
    donut.title = "Distribuição de Status dos Lotes"
    donut.style = 10
    
    data_donut = Reference(ws_resumo, min_col=2, min_row=3, max_row=6)
    labels_donut = Reference(ws_resumo, min_col=1, min_row=3, max_row=6)
    donut.add_data(data_donut, titles_from_data=False)
    donut.set_categories(labels_donut)
    donut.width = 15
    donut.height = 10
    
    ws_resumo.add_chart(donut, "E2")

    # B) Gráfico de Linha (Evolução diária de problemas)
    df_problemas = df_todos[df_todos["Classificação"].isin(["Divergência", "Ambíguo"])]
    
    agrupado = df_problemas.groupby("Data").size()
    evolucao_diaria = agrupado.reset_index(name="Total_Problemas")
    evolucao_diaria = evolucao_diaria.sort_values("Data")

    ws_resumo["G1"] = "Data"
    ws_resumo["H1"] = "Divergências + Ambíguos"
    
    for idx, row in enumerate(evolucao_diaria.itertuples(), start=2):
        ws_resumo.cell(row=idx, column=7, value=str(row.Data))
        ws_resumo.cell(row=idx, column=8, value=row.Total_Problemas)

    max_row_ev = max(2, len(evolucao_diaria) + 1)

    line = LineChart()
    line.title = "Evolução Diária de Problemas (Divergências + Ambíguos)"
    line.style = 13
    line.y_axis.title = "Quantidade"
    line.x_axis.title = "Data"

    data_line = Reference(ws_resumo, min_col=8, min_row=1, max_row=max_row_ev)
    labels_line = Reference(ws_resumo, min_col=7, min_row=2, max_row=max_row_ev)
    line.add_data(data_line, titles_from_data=True)
    line.set_categories(labels_line)
    line.width = 18
    line.height = 10

    ws_resumo.add_chart(line, "E17")

    wb.save(caminho_saida)
    wb.close()
    return caminho_saida


def gerar_resumo_executivo(
    indicadores: OperationalIndicators,
    caminho_saida=ARQUIVO_RESUMO_EXECUTIVO,
) -> Path:
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    texto = f"""# Resumo Executivo da Conferência de Lotes

## Visão Geral

Foram processados {indicadores.total_registros} {_plural(indicadores.total_registros, "registro", "registros")} de inspeção. O resultado consolidado mostra {indicadores.registros_validos} {_plural(indicadores.registros_validos, "registro válido", "registros válidos")}, {indicadores.divergencias} {_plural(indicadores.divergencias, "divergência", "divergências")}, {indicadores.ambiguos} {_plural(indicadores.ambiguos, "caso ambíguo", "casos ambíguos")} e {indicadores.erros_entrada} {_plural(indicadores.erros_entrada, "erro de entrada", "erros de entrada")}.

## Indicadores Principais

- Registros válidos: {indicadores.registros_validos} ({_fmt_pct(indicadores.percentual_validos)})
- Divergências: {indicadores.divergencias} ({_fmt_pct(indicadores.percentual_divergencias)})
- Ambíguos: {indicadores.ambiguos} ({_fmt_pct(indicadores.percentual_ambiguos)})
- Erros de entrada: {indicadores.erros_entrada} ({_fmt_pct(indicadores.percentual_erros_entrada)})
- Taxa de qualidade da entrada: {_fmt_pct(indicadores.taxa_qualidade_entrada)}
- Taxa de revisão humana: {_fmt_pct(indicadores.taxa_revisao_humana)}
- Taxa de retrabalho: {_fmt_pct(indicadores.taxa_retrabalho)}

## Destaque

A regra mais acionada foi {indicadores.regra_mais_acionada} ({indicadores.regra_mais_acionada_nome}), com {indicadores.regra_mais_acionada_quantidade} {_plural(indicadores.regra_mais_acionada_quantidade, "ocorrência", "ocorrências")}. Esse ponto indica o principal gargalo operacional observado na rodada.

## Ganho Estimado de Tempo

Premissas usadas: {indicadores.tempo_manual_minutos_por_registro:.0f} {_plural(int(indicadores.tempo_manual_minutos_por_registro), "minuto", "minutos")} por registro em conferência manual e {indicadores.tempo_automatizado_minutos_por_registro:.0f} {_plural(int(indicadores.tempo_automatizado_minutos_por_registro), "minuto", "minutos")} por registro no fluxo automatizado. Com essas premissas, o ganho estimado é de {indicadores.ganho_estimado_minutos:.0f} minutos, ou {_fmt_decimal(indicadores.ganho_estimado_horas)} horas.

## Observação

O ganho de tempo é uma estimativa didática para apoiar a avaliação do exercício. Para virar uma métrica real de produção, seria necessário medir tempos executados em ambiente produtivo, com amostra controlada e histórico comparável.
"""

    caminho_saida.write_text(texto, encoding="utf-8")
    return caminho_saida


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------

def main(
    caminho_entrada=ARQUIVO_ENTRADA,
    caminho_saida=ARQUIVO_SAIDA,
    caminho_resumo=ARQUIVO_RESUMO_EXECUTIVO,
    decisoes_ml: list[dict] | None = None,
    ml_client=None,
):
    registros = consolidar_e_validar(caminho_entrada)
    indicadores = consolidar_indicadores(registros)
    if decisoes_ml is None:
        ml_client = ml_client or MLClient(
            ML_API_URL,
            timeout=ML_TIMEOUT_SECONDS,
            max_failures=ML_MAX_FAILURES,
        )
        decisoes_ml = classificar_registros_ambiguos_com_ml(
            registros,
            ml_client,
        )
    caminho_gerado = gerar_relatorio_excel(
        registros,
        caminho_saida,
        indicadores,
        decisoes_ml,
    )
    caminho_resumo_gerado = gerar_resumo_executivo(indicadores, caminho_resumo)

    total = len(registros)
    contagem = Counter(r.classificacao for r in registros)
    print(f"Registros processados: {total}")
    for categoria in ["Válido", "Divergência", "Ambíguo", "Erro de Entrada"]:
        print(f"  {categoria}: {contagem.get(categoria, 0)}")
    print(
        "Decisões ML: "
        f"{len(decisoes_ml)} de {contagem.get('Ambíguo', 0)} casos ambíguos"
    )
    print(
        "Regra mais acionada: "
        f"{indicadores.regra_mais_acionada} "
        f"({indicadores.regra_mais_acionada_quantidade})"
    )
    print(f"Relatório gerado em: {caminho_gerado.resolve()}")
    print(f"Resumo executivo gerado em: {caminho_resumo_gerado.resolve()}")

    return registros


if __name__ == "__main__":
    main()
